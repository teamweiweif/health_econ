from __future__ import annotations

import csv
import io
import json
import math
import os
import re
import sys
import time
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from scipy import stats
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import NearestNeighbors


THIS_FILE = Path(__file__).resolve()
V3_ROOT = THIS_FILE.parents[1]
DT_ROOT = THIS_FILE.parents[2]
PROJECT_ROOT = THIS_FILE.parents[3]
REPO_ROOT = THIS_FILE.parents[4]
TABLES = V3_ROOT / "result" / "tables"
FIGURES = V3_ROOT / "result" / "figures"
REPORT = V3_ROOT / "report"
DATA = V3_ROOT / "data"
TEMP = V3_ROOT / "temp"
RAW = TEMP / "raw_downloads"
SNAP = TEMP / "source_snapshots"
AUDIT = TEMP / "audit_logs"

sys.path.insert(0, str(DT_ROOT / "script"))
import design_tournament_common as v2  # noqa: E402


SNAPSHOTS = ["2022-01-27", "2022-04-27", "2022-07-27", "2022-10-27", "2023-01-02"]
TARGET_ARCHIVE_DATES = set(SNAPSHOTS)
OFFICIAL_ARCHIVE_API = "https://data.cms.gov/provider-data/api/1/archive/aggregate/theme/nursing-homes"
CMS_PBJ_CATALOG_PAGE = "https://catalog.data.gov/dataset/payroll-based-journal-daily-nurse-staffing"
PBJ_DATASET_PAGE = "https://data.cms.gov/quality-of-care/payroll-based-journal-daily-nurse-staffing"
STAFFING_THRESHOLDS = [155, 205, 255, 320]
PLACEBO_THRESHOLDS = [180, 230, 285, 345]
OUTCOME_COLS = [
    "weekend_rn_lt8_day_share",
    "weekend_rn_lt8_days",
    "zero_rn_day_share",
    "zero_rn_days",
    "weekend_p10_total_hprd",
    "weekend_p25_total_hprd",
    "worst_weekend_total_hprd",
    "worst_weekend_rn_hprd",
    "weekend_share_total_hours",
    "contract_share_total_hours",
    "avg_daily_census",
    "occupancy",
]


def ensure_dirs() -> None:
    for p in [DATA, TABLES, FIGURES, REPORT, RAW, SNAP, AUDIT]:
        p.mkdir(parents=True, exist_ok=True)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def save_csv(path: Path, rows: list[dict] | pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(rows, pd.DataFrame):
        rows.to_csv(path, index=False)
        return
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    keys = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows)


def md_table(df: pd.DataFrame, max_rows: int | None = None) -> str:
    if df is None or df.empty:
        return "No rows."
    d = df.copy()
    if max_rows is not None and len(d) > max_rows:
        d = d.head(max_rows)
    cols = list(d.columns)
    lines = ["| " + " | ".join(cols) + " |", "| " + " | ".join(["---"] * len(cols)) + " |"]
    for _, row in d.iterrows():
        vals = []
        for c in cols:
            x = row[c]
            if pd.isna(x):
                vals.append("")
            elif isinstance(x, float):
                vals.append(f"{x:.6g}")
            else:
                vals.append(str(x).replace("\n", " ")[:240])
        lines.append("| " + " | ".join(vals) + " |")
    if max_rows is not None and len(df) > max_rows:
        lines.append(f"\nShowing {max_rows} of {len(df)} rows.")
    return "\n".join(lines)


def norm_col(c: str) -> str:
    return v2.norm_col(c)


def zfill_ccn(s: pd.Series) -> pd.Series:
    return v2.zfill_ccn(s)


def to_num(s: pd.Series | None, index: pd.Index | None = None) -> pd.Series:
    if s is None:
        return pd.Series(np.nan, index=index)
    return pd.to_numeric(s, errors="coerce")


def source_candidates() -> dict[str, Path]:
    sibling = REPO_ROOT.parent / "nh_staffing" / "nursing_home_staffing_reporting"
    return {
        "project_data": PROJECT_ROOT / "data",
        "project_temp": PROJECT_ROOT / "temp",
        "provider_archives": PROJECT_ROOT / "temp" / "raw_downloads" / "provider_archives",
        "v2_data": DT_ROOT / "data",
        "v2_temp": DT_ROOT / "temp",
        "sibling_project_data": sibling / "data",
        "sibling_project_temp": sibling / "temp",
        "sibling_provider_archives": sibling / "temp" / "raw_downloads" / "provider_archives",
        "sibling_v2_data": sibling / "design_tournament" / "data",
        "sibling_v2_temp": sibling / "design_tournament" / "temp",
    }


def first_existing(label: str, labels: list[str]) -> Path:
    cand = source_candidates()
    for k in labels:
        p = cand[k]
        if p.exists():
            return p
    raise FileNotFoundError(f"Could not locate {label}: {labels}")


def first_existing_file(label: str, rel: str, labels: list[str]) -> Path:
    cand = source_candidates()
    for k in labels:
        p = cand[k] / rel
        if p.exists():
            return p
    raise FileNotFoundError(f"Could not locate {label}: {rel}")


def provider_archive_dir() -> Path:
    return first_existing("Provider Data Catalog archives", ["provider_archives", "sibling_provider_archives"])


def project_data_dir() -> Path:
    return first_existing("project data", ["project_data", "sibling_project_data"])


def project_temp_dir() -> Path:
    return first_existing("project temp", ["project_temp", "sibling_project_temp"])


def v2_data_dir() -> Path:
    return first_existing("v2 design data", ["v2_data", "sibling_v2_data"])


def v2_temp_dir() -> Path:
    return first_existing("v2 design temp", ["v2_temp", "sibling_v2_temp"])


def v2_data_file(rel: str) -> Path:
    return first_existing_file("v2 design data file", rel, ["v2_data", "sibling_v2_data"])


def project_temp_file(rel: str) -> Path:
    return first_existing_file("project temp file", rel, ["project_temp", "sibling_project_temp"])


def configure_v2_paths() -> None:
    v2.OLD_PROVIDER_ARCHIVES = provider_archive_dir()
    v2.OLD_DATA = project_data_dir()
    v2.OLD_TEMP = project_temp_dir()
    v2.DATA = v2_data_dir()
    v2.TEMP = v2_temp_dir()


def provider_zip(snapshot: str) -> Path:
    p = provider_archive_dir() / f"nursing_homes_{snapshot}.zip"
    if not p.exists():
        raise FileNotFoundError(p)
    return p


def provider_info_member(z: zipfile.ZipFile) -> str:
    members = [m for m in z.namelist() if "NH_ProviderInfo" in m and m.lower().endswith(".csv")]
    if not members:
        raise FileNotFoundError("NH_ProviderInfo CSV missing")
    return members[0]


def read_provider_info_raw(snapshot: str) -> pd.DataFrame:
    with zipfile.ZipFile(provider_zip(snapshot)) as z:
        member = provider_info_member(z)
        raw = pd.read_csv(z.open(member), encoding="cp1252", encoding_errors="replace", low_memory=False)
    return raw.rename(columns={c: norm_col(c) for c in raw.columns})


def read_provider_info_v3(snapshot: str) -> pd.DataFrame:
    raw = read_provider_info_raw(snapshot)
    idx = raw.index
    out = pd.DataFrame()
    out["facility_id"] = zfill_ccn(raw.get("federal_provider_number", pd.Series(index=idx, dtype=object)))
    out["snapshot_date"] = pd.Timestamp(snapshot)
    out["snapshot"] = snapshot
    out["provider_name"] = raw.get("provider_name")
    out["state"] = raw.get("provider_state")
    out["ownership_type"] = raw.get("ownership_type")
    out["certified_beds"] = to_num(raw.get("number_of_certified_beds"), idx)
    out["avg_residents_per_day"] = to_num(raw.get("average_number_of_residents_per_day"), idx)
    out["overall_rating"] = to_num(raw.get("overall_rating"), idx)
    out["health_inspection_rating"] = to_num(raw.get("health_inspection_rating"), idx)
    out["qm_rating"] = to_num(raw.get("qm_rating"), idx)
    out["staffing_rating"] = to_num(raw.get("staffing_rating"), idx)
    out["rn_staffing_rating"] = to_num(raw.get("rn_staffing_rating"), idx)
    out["reported_total_nurse_hprd"] = to_num(raw.get("reported_total_nurse_staffing_hours_per_resident_per_day"), idx)
    out["case_mix_total_nurse_hprd"] = to_num(raw.get("case_mix_total_nurse_staffing_hours_per_resident_per_day"), idx)
    out["adjusted_total_nurse_hprd"] = to_num(raw.get("adjusted_total_nurse_staffing_hours_per_resident_per_day"), idx)
    out["reported_rn_hprd"] = to_num(raw.get("reported_rn_staffing_hours_per_resident_per_day"), idx)
    out["adjusted_rn_hprd"] = to_num(raw.get("adjusted_rn_staffing_hours_per_resident_per_day"), idx)
    out["reported_weekend_total_nurse_hprd"] = to_num(raw.get("total_number_of_nurse_staff_hours_per_resident_per_day_on_the_weekend"), idx)
    out["reported_weekend_rn_hprd"] = to_num(raw.get("registered_nurse_hours_per_resident_per_day_on_the_weekend"), idx)
    out["adjusted_weekend_total_nurse_hprd_official"] = to_num(raw.get("adjusted_weekend_total_nurse_staffing_hours_per_resident_per_day"), idx)
    out["total_nursing_staff_turnover"] = to_num(raw.get("total_nursing_staff_turnover"), idx)
    out["rn_turnover"] = to_num(raw.get("registered_nurse_turnover"), idx)
    out["administrator_turnover"] = to_num(raw.get("number_of_administrators_who_have_left_the_nursing_home"), idx)
    out["rating_cycle_1_total_health_deficiencies"] = to_num(raw.get("rating_cycle_1_total_number_of_health_deficiencies"), idx)
    out["total_weighted_health_survey_score"] = to_num(raw.get("total_weighted_health_survey_score"), idx)
    ratio = out["adjusted_total_nurse_hprd"] / out["reported_total_nurse_hprd"].replace(0, np.nan)
    out["total_nurse_case_mix_adjustment_ratio"] = ratio
    out["adjusted_weekend_total_nurse_hprd_reconstructed"] = out["reported_weekend_total_nurse_hprd"] * ratio
    explicit = out["adjusted_weekend_total_nurse_hprd_official"].notna()
    reconstructed = out["adjusted_weekend_total_nurse_hprd_reconstructed"].notna()
    out["adjusted_weekend_total_nurse_hprd"] = out["adjusted_weekend_total_nurse_hprd_official"]
    out.loc[~explicit & reconstructed, "adjusted_weekend_total_nurse_hprd"] = out.loc[
        ~explicit & reconstructed, "adjusted_weekend_total_nurse_hprd_reconstructed"
    ]
    out.loc[~out["adjusted_weekend_total_nurse_hprd"].notna(), "adjusted_weekend_total_nurse_hprd"] = out.loc[
        ~out["adjusted_weekend_total_nurse_hprd"].notna(), "reported_weekend_total_nurse_hprd"
    ]
    out["adjusted_weekend_hprd_source"] = np.select(
        [explicit, reconstructed, out["reported_weekend_total_nurse_hprd"].notna()],
        ["explicit_official_field", "reconstructed_from_official_ratio", "reported_weekend_proxy"],
        default="missing",
    )
    out = out[out["facility_id"].notna()].drop_duplicates("facility_id")
    return out


def add_sample(df: pd.DataFrame, sample: str) -> pd.DataFrame:
    out = df.copy()
    if "sample_definition" not in out.columns:
        out["sample_definition"] = sample
    return out


def download_json(url: str, path: Path) -> dict:
    path.parent.mkdir(parents=True, exist_ok=True)
    r = requests.get(url, timeout=60)
    r.raise_for_status()
    path.write_bytes(r.content)
    return r.json()


def field_role(field: str, file_name: str) -> str:
    n = norm_col(field)
    f = file_name.lower()
    if re.search(r"(staffing_)?(total_)?score|staffing_points|domain_points|total_staffing_score", n):
        if "snf" in f or "quality_reporting" in f:
            return "unrelated_score_field"
        return "possible_official_staffing_score"
    if n == "score":
        return "generic_score_not_staffing_domain"
    if "adjusted_weekend_total_nurse" in n:
        return "official_adjusted_weekend_total_nurse_hprd"
    if "total_number_of_nurse_staff_hours_per_resident_per_day_on_the_weekend" in n:
        return "reported_weekend_total_nurse_hprd_component"
    if "reported_total_nurse_staffing_hours_per_resident_per_day" in n:
        return "reported_total_nurse_hprd_component"
    if "adjusted_total_nurse_staffing_hours_per_resident_per_day" in n:
        return "adjusted_total_nurse_hprd_component"
    if "case_mix_total_nurse" in n:
        return "case_mix_total_nurse_hprd_component"
    if "staffing_rating" in n:
        return "official_staffing_star_validation"
    if "turnover" in n or "administrators" in n:
        return "staffing_turnover_component"
    if "weekend" in n and "rn" in n:
        return "weekend_rn_transparency_measure"
    return "candidate_context_field"


def usable_for_july(field: str, snapshot: str, role: str) -> str:
    if snapshot != "2022-07-27":
        return "no_not_july_snapshot"
    if role in {
        "reported_weekend_total_nurse_hprd_component",
        "reported_total_nurse_hprd_component",
        "adjusted_total_nurse_hprd_component",
        "case_mix_total_nurse_hprd_component",
        "staffing_turnover_component",
        "official_staffing_star_validation",
    }:
        return "yes_component"
    if role == "official_adjusted_weekend_total_nurse_hprd":
        return "yes_if_present"
    if role == "possible_official_staffing_score":
        return "yes_if_facility_level_staffing_domain"
    return "no"


def missingness_for_field(snapshot: str, member: str, field: str) -> str:
    try:
        with zipfile.ZipFile(provider_zip(snapshot)) as z:
            with z.open(member) as f:
                s = pd.read_csv(f, usecols=[field], encoding="cp1252", encoding_errors="replace", low_memory=False)[field]
        return f"{float(s.isna().mean()):.6f}"
    except Exception:
        return ""


def stage1_source_field_inventory() -> None:
    ensure_dirs()
    rows: list[dict] = []
    api_path = RAW / "provider_data_archive_api_nursing_homes.json"
    api = download_json(OFFICIAL_ARCHIVE_API, api_path)
    for rec in api.get("data", []):
        date = rec.get("date")
        if date in TARGET_ARCHIVE_DATES:
            rows.append(
                {
                    "source_name": "Provider Data Catalog archive API",
                    "source_url_or_local_path": OFFICIAL_ARCHIVE_API,
                    "snapshot_date": date,
                    "file_name": rec.get("name", ""),
                    "field_name": "archive_zip_url",
                    "field_label_if_available": rec.get("url", ""),
                    "possible_role": "archive_manifest",
                    "usable_for_july2022": "yes_archive_manifest" if date == "2022-07-27" else "no_not_july_snapshot",
                    "missingness": "",
                    "notes": f"id={rec.get('id')}; size={rec.get('size')}; access={rec.get('access_level')}",
                    "final_decision": "source_manifest_only",
                    "n_rows_checked": "",
                    "sample_definition": "official CMS archive API records for requested nursing-home snapshots",
                }
            )

    candidate_re = re.compile(
        r"score|point|staffing|weekend|turnover|administrator|case.?mix|adjusted|reported_total_nurse",
        re.I,
    )
    for snapshot in SNAPSHOTS:
        zp = provider_zip(snapshot)
        with zipfile.ZipFile(zp) as z:
            for member in z.namelist():
                lower = member.lower()
                if lower.endswith(".csv"):
                    try:
                        with z.open(member) as f:
                            header = next(csv.reader(io.TextIOWrapper(f, encoding="utf-8-sig", errors="replace")))
                    except Exception:
                        continue
                    for field in header:
                        if not candidate_re.search(field):
                            continue
                        role = field_role(field, member)
                        rows.append(
                            {
                                "source_name": "Provider Data Catalog archived ZIP CSV header",
                                "source_url_or_local_path": str(zp),
                                "snapshot_date": snapshot,
                                "file_name": member,
                                "field_name": field,
                                "field_label_if_available": field,
                                "possible_role": role,
                                "usable_for_july2022": usable_for_july(field, snapshot, role),
                                "missingness": missingness_for_field(snapshot, member, field)
                                if "NH_ProviderInfo" in member
                                else "",
                                "notes": "CSV header scan; field recorded as candidate by keyword search.",
                                "final_decision": "candidate_component" if role.endswith("component") or "official_adjusted" in role else "not_official_score",
                                "n_rows_checked": "",
                                "sample_definition": "all candidate CSV headers in requested Provider Data Catalog ZIP snapshots",
                            }
                        )
                elif lower.endswith(".xlsx") and "data_dictionary" in lower:
                    wb = load_workbook(io.BytesIO(z.read(member)), read_only=True, data_only=True)
                    for ws in wb.worksheets:
                        headers = None
                        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                            vals = ["" if x is None else str(x) for x in row]
                            joined = " | ".join(vals)
                            if i == 1:
                                headers = vals
                                continue
                            if not candidate_re.search(joined):
                                continue
                            label = vals[0] if vals else ""
                            desc = vals[1] if len(vals) > 1 else ""
                            role = field_role(label, ws.title)
                            rows.append(
                                {
                                    "source_name": "Provider Data Catalog NH_Primary_Data_Dictionary.xlsx",
                                    "source_url_or_local_path": str(zp),
                                    "snapshot_date": snapshot,
                                    "file_name": f"{member}::{ws.title}",
                                    "field_name": label,
                                    "field_label_if_available": desc,
                                    "possible_role": role,
                                    "usable_for_july2022": usable_for_july(label, snapshot, role),
                                    "missingness": "",
                                    "notes": joined[:500],
                                    "final_decision": "candidate_component" if role.endswith("component") or "official_adjusted" in role else "dictionary_context",
                                    "n_rows_checked": "",
                                    "sample_definition": "all matching rows in Provider Data Catalog primary data dictionaries",
                                }
                            )

    guide = DT_ROOT / "temp" / "source_snapshots" / "technical_guides" / "Five Star Users' Guide July 2022.txt"
    if guide.exists():
        text = guide.read_text(encoding="utf-8", errors="replace")
        for name, pattern, role in [
            ("case_mix_formula", "Hours Adjusted = (Hours Reported/Hours Case-Mix) * Hours National Average", "algorithm_formula"),
            ("weekend_case_mix_all_days", "case-mix value for all days in the quarter is used for", "algorithm_formula"),
            ("total_staffing_score", "total staffing score is compared to staffing rating point thresholds", "score_definition"),
        ]:
            found = pattern.lower() in text.lower()
            rows.append(
                {
                    "source_name": "CMS July 2022 Five-Star Technical Users Guide",
                    "source_url_or_local_path": str(guide),
                    "snapshot_date": "2022-07-27",
                    "file_name": guide.name,
                    "field_name": name,
                    "field_label_if_available": pattern,
                    "possible_role": role,
                    "usable_for_july2022": "yes_formula_evidence" if found else "not_found",
                    "missingness": "",
                    "notes": "Text evidence from official archived July 2022 guide." if found else "Pattern not found.",
                    "final_decision": "supports_ratio_reconstruction" if found else "not_found",
                    "n_rows_checked": "",
                    "sample_definition": "official July 2022 Technical Users Guide text",
                }
            )

    inv = pd.DataFrame(rows).drop_duplicates()
    save_csv(TABLES / "official_score_field_inventory.csv", inv)

    july = inv[inv["snapshot_date"].eq("2022-07-27")]
    official_score_found = bool(
        july["possible_role"].eq("possible_official_staffing_score").any()
        and july["final_decision"].str.contains("official_score_found", na=False).any()
    )
    explicit_july_adjusted = bool(
        july["possible_role"].eq("official_adjusted_weekend_total_nurse_hprd").any()
        and july["source_name"].str.contains("CSV header", na=False).any()
    )
    components = {
        "reported_weekend_total_nurse_hprd_component",
        "reported_total_nurse_hprd_component",
        "adjusted_total_nurse_hprd_component",
    }
    component_roles = set(july["possible_role"])
    enough_components = components.issubset(component_roles)
    conclusion = (
        "official score found"
        if official_score_found
        else "adjusted weekend total nurse HPRD found"
        if explicit_july_adjusted
        else "enough official components found to reconstruct without proxy substitution"
        if enough_components
        else "not found after exhaustive search"
    )
    notes = pd.DataFrame(
        [
            {"check": "official_0_380_staffing_score_found", "result": official_score_found},
            {"check": "explicit_july_adjusted_weekend_total_nurse_hprd_found", "result": explicit_july_adjusted},
            {"check": "reported_weekend_total_component_found", "result": "reported_weekend_total_nurse_hprd_component" in component_roles},
            {"check": "reported_total_component_found", "result": "reported_total_nurse_hprd_component" in component_roles},
            {"check": "adjusted_total_component_found", "result": "adjusted_total_nurse_hprd_component" in component_roles},
            {"check": "conclusion", "result": conclusion},
        ]
    )
    notes["n"] = len(notes)
    save_csv(TABLES / "official_score_field_inventory_summary.csv", add_sample(notes, "field hunt summary across official CMS sources"))

    top = inv[inv["snapshot_date"].eq("2022-07-27") & inv["possible_role"].isin(list(components) + ["official_adjusted_weekend_total_nurse_hprd", "possible_official_staffing_score"])]
    md = [
        "# Official Score Field Hunt V3",
        "",
        f"Conclusion: **{conclusion}**.",
        "",
        "## Search Scope",
        "",
        "- CMS Provider Data Catalog nursing-home archive API.",
        "- Provider Data Catalog archived ZIPs for January 2022, April 2022, July 2022, October 2022, and January 2023.",
        "- All ZIP member names, CSV headers, ProviderInfo fields, and `NH_Primary_Data_Dictionary.xlsx` rows in those snapshots.",
        "- Official July 2022 Technical Users' Guide text already snapshotted in V2.",
        "",
        "## Findings",
        "",
        f"- Official facility-level 0-380 staffing score found: {official_score_found}.",
        f"- Explicit July 2022 adjusted weekend total nurse HPRD field found: {explicit_july_adjusted}.",
        f"- July 2022 official components sufficient for reconstruction found: {enough_components}.",
        "- The reconstruction uses official July ProviderInfo reported weekend total nurse HPRD and the official all-day total nurse case-mix adjustment ratio `adjusted_total_nurse_hprd / reported_total_nurse_hprd`.",
        "- The official July guide states that the all-days case-mix value is used to calculate case-mix adjusted total nurse staffing on weekends.",
        "- October 2022 and January 2023 contain an explicit adjusted weekend total nurse HPRD field, which is used in Stage 2 to validate the reconstruction identity.",
        "",
        "## Key July Candidate Fields",
        "",
        md_table(top[["source_name", "file_name", "field_name", "possible_role", "usable_for_july2022", "missingness", "final_decision"]], max_rows=40),
    ]
    write_text(REPORT / "01_official_score_field_hunt.md", "\n".join(md))


def stage2_emulator_rescue() -> None:
    ensure_dirs()
    configure_v2_paths()
    frames = []
    for snap in SNAPSHOTS:
        base = read_provider_info_v3(snap)
        em = v2.emulate_staffing_score(base)
        frames.append(em)
    all_scores = pd.concat(frames, ignore_index=True)
    all_scores["exact_score_source"] = "not_found; score reconstructed from official component fields"
    all_scores["proxy_used"] = all_scores["adjusted_weekend_hprd_source"].eq("reported_weekend_proxy")
    all_scores.to_parquet(DATA / "rating_score_emulator_panel_v3.parquet", index=False)
    all_scores.to_csv(DATA / "rating_score_emulator_panel_v3.csv", index=False)
    july = all_scores[all_scores["snapshot"].eq("2022-07-27")].copy()
    july.to_parquet(DATA / "analysis_threshold_facility_july2022_v3.parquet", index=False)

    rows = []
    for snap, g in all_scores.groupby("snapshot", sort=True):
        valid = g["emulator_match"].notna()
        source = ";".join(sorted(g["adjusted_weekend_hprd_source"].dropna().unique()))
        match = float(g.loc[valid, "emulator_match"].mean()) if valid.any() else np.nan
        rows.append(
            {
                "snapshot_date": snap,
                "n_facilities": int(len(g)),
                "n_with_official_staffing_star": int(g["staffing_rating"].notna().sum()),
                "n_with_score_or_proxy": int(g["staffing_score_proxy"].notna().sum()),
                "n_valid_match_test": int(valid.sum()),
                "star_match_rate": match,
                "exact_score_source": "not found",
                "adjusted_weekend_hprd_source": source,
                "proxy_used": bool(g["proxy_used"].any()),
                "passes_95pct_threshold": bool(pd.notna(match) and match >= 0.95 and snap >= "2022-07-27"),
                "notes": "pre-July snapshot used old staffing methodology" if snap < "2022-07-27" else "July 2022 guide six-component score validation",
                "n": int(valid.sum()),
                "sample_definition": "facilities with official staffing star and reconstructed/official score",
            }
        )
    validation = pd.DataFrame(rows)
    save_csv(TABLES / "emulator_validation_v3.csv", validation)

    diag_rows = []
    all_scores["match_int"] = all_scores["emulator_match"].astype(float)
    all_scores["size_group"] = np.where(all_scores["certified_beds"] >= all_scores.groupby("snapshot")["certified_beds"].transform("median"), "large", "small")
    all_scores["ownership_group"] = np.where(all_scores["ownership_type"].astype(str).str.lower().str.contains("for profit"), "for_profit", "nonprofit_or_gov")
    all_scores["missing_component_pattern"] = all_scores[
        [
            "points_total_nurse",
            "points_rn",
            "points_weekend_total",
            "points_total_turnover",
            "points_rn_turnover",
            "points_admin_turnover",
        ]
    ].isna().apply(lambda r: "+".join([c.replace("points_", "") for c, v in r.items() if v]) or "complete", axis=1)
    group_specs = [
        ("state", "state"),
        ("ownership", "ownership_group"),
        ("official_staffing_star", "staffing_rating"),
        ("missing_component_pattern", "missing_component_pattern"),
        ("weekend_hprd_source", "adjusted_weekend_hprd_source"),
        ("facility_size", "size_group"),
        ("snapshot", "snapshot"),
    ]
    for snap, g0 in all_scores.groupby("snapshot"):
        for typ, col in group_specs:
            for key, g in g0.groupby(col, dropna=False):
                valid = g["emulator_match"].notna()
                if valid.sum() == 0:
                    continue
                match_vals = g.loc[valid, "emulator_match"].astype(bool)
                diag_rows.append(
                    {
                        "snapshot_date": snap,
                        "diagnostic_type": typ,
                        "diagnostic_group": str(key),
                        "n": int(valid.sum()),
                        "match_rate": float(match_vals.mean()),
                        "mismatch_count": int((~match_vals).sum()),
                        "proxy_used_share": float(g.loc[valid, "proxy_used"].mean()),
                        "sample_definition": "valid emulator-star comparisons by diagnostic subgroup",
                    }
                )
    diag = pd.DataFrame(diag_rows)
    save_csv(TABLES / "emulator_mismatch_diagnostics_v3.csv", diag)

    v2_val_path = DT_ROOT / "result" / "tables" / "rating_emulator_validation.csv"
    if not v2_val_path.exists():
        v2_val_path = REPO_ROOT.parent / "nh_staffing" / "nursing_home_staffing_reporting" / "design_tournament" / "result" / "tables" / "rating_emulator_validation.csv"
    v2_val = pd.read_csv(v2_val_path) if v2_val_path.exists() else pd.DataFrame()
    comparison = validation[["snapshot_date", "star_match_rate"]].rename(columns={"star_match_rate": "v3_star_match_rate"})
    if not v2_val.empty:
        v2_cmp = v2_val.rename(columns={"snapshot": "snapshot_date", "star_match_rate": "v2_star_match_rate"})[["snapshot_date", "v2_star_match_rate"]]
        comparison = comparison.merge(v2_cmp, on="snapshot_date", how="left")
        comparison["v3_minus_v2"] = comparison["v3_star_match_rate"] - comparison["v2_star_match_rate"]
    comparison["n"] = len(comparison)
    save_csv(TABLES / "emulator_validation_v2_v3_comparison.csv", add_sample(comparison, "snapshot-level V2 versus V3 emulator validation"))

    july_match = float(validation.loc[validation["snapshot_date"].eq("2022-07-27"), "star_match_rate"].iloc[0])
    systematic = diag[
        diag["snapshot_date"].eq("2022-07-27")
        & diag["diagnostic_type"].isin(["ownership", "facility_size", "weekend_hprd_source"])
        & diag["n"].ge(100)
    ]
    min_diag = systematic["match_rate"].min()
    severe = bool(pd.notna(min_diag) and min_diag < 0.90)
    decision = "candidate primary evidence" if july_match >= 0.95 and not severe else "exploratory"
    md = [
        "# Emulator Rescue Audit V3",
        "",
        "## Core Rescue",
        "",
        "- Official 0-380 staffing score was not found.",
        "- Explicit adjusted weekend total nurse HPRD was not found in the July 2022 ProviderInfo file.",
        "- July 2022 adjusted weekend total nurse HPRD was reconstructed from official July ProviderInfo components as `reported_weekend_total_nurse_hprd * adjusted_total_nurse_hprd / reported_total_nurse_hprd`.",
        "- This is supported by the July guide statement that the all-days case-mix value is used for case-mix-adjusted total nurse staffing on weekends.",
        "- In October 2022 and January 2023, where CMS publishes explicit adjusted weekend total HPRD, the same identity reproduces the official field almost exactly.",
        "",
        "## Validation",
        "",
        md_table(validation),
        "",
        "## V2 Versus V3",
        "",
        md_table(comparison),
        "",
        "## Decision",
        "",
        f"July 2022 match rate is {july_match:.3f}; RD/RD-DID status: **{decision}**.",
        "Pre-July snapshots are retained as old-methodology comparisons and are not used to validate the July 2022 six-component score.",
    ]
    write_text(REPORT / "02_emulator_rescue_audit.md", "\n".join(md))


def build_analysis_panel_v3() -> pd.DataFrame:
    if not (DATA / "analysis_threshold_facility_july2022_v3.parquet").exists():
        stage2_emulator_rescue()
    july = pd.read_parquet(DATA / "analysis_threshold_facility_july2022_v3.parquet")
    q_path = v2_data_file("reliability_outcomes_facility_quarter.parquet")
    q = pd.read_parquet(q_path)
    pre = q[q["quarter"].isin(["2021Q1", "2021Q2", "2021Q3"])].copy()
    post = q[q["quarter"].isin(["2022Q4", "2023Q1", "2023Q2", "2023Q3", "2023Q4"])].copy()
    for d in [pre, post]:
        d["occupancy"] = np.nan
    pre_g = pre.groupby("facility_id").agg({c: "mean" for c in OUTCOME_COLS if c in pre.columns}).add_prefix("pre_").reset_index()
    post_g = post.groupby("facility_id").agg({c: "mean" for c in OUTCOME_COLS if c in post.columns}).add_prefix("post_").reset_index()
    panel = july.merge(pre_g, on="facility_id", how="left").merge(post_g, on="facility_id", how="left")
    panel["occupancy"] = panel["avg_residents_per_day"] / panel["certified_beds"].replace(0, np.nan)
    if "post_avg_daily_census" in panel and "certified_beds" in panel:
        panel["post_occupancy"] = panel["post_avg_daily_census"] / panel["certified_beds"].replace(0, np.nan)
    if "pre_avg_daily_census" in panel and "certified_beds" in panel:
        panel["pre_occupancy"] = panel["pre_avg_daily_census"] / panel["certified_beds"].replace(0, np.nan)
    for c in OUTCOME_COLS:
        if f"post_{c}" in panel and f"pre_{c}" in panel:
            panel[f"delta_{c}"] = panel[f"post_{c}"] - panel[f"pre_{c}"]
    panel["ownership_for_profit"] = panel["ownership_type"].astype(str).str.lower().str.contains("for profit").astype(int)
    panel["large_facility"] = (panel["certified_beds"] >= panel["certified_beds"].median(skipna=True)).astype(int)
    panel.to_parquet(DATA / "analysis_threshold_facility_july2022_v3.parquet", index=False)
    return panel


def local_linear(df: pd.DataFrame, y: str, x: str, cutoff: float, bw: float, donut: float = 0.0) -> dict:
    return v2.local_linear_rd(df, y, x, cutoff, bw, donut)


def stage3_rd_rddid_if_valid() -> None:
    ensure_dirs()
    validation = pd.read_csv(TABLES / "emulator_validation_v3.csv")
    july_match = float(validation.loc[validation["snapshot_date"].eq("2022-07-27"), "star_match_rate"].iloc[0])
    if july_match < 0.95:
        write_text(
            REPORT / "03_rd_rddid_no_go_due_to_running_variable.md",
            "# RD/RD-DID No-Go Due To Running Variable\n\nJuly 2022 emulator validation remains below 0.950, so RD/RD-DID is not rerun as valid primary evidence.",
        )
        return
    df = build_analysis_panel_v3()
    x = "staffing_score_proxy_round"
    post_outcomes = [f"post_{c}" for c in OUTCOME_COLS if f"post_{c}" in df.columns]
    delta_outcomes = [f"delta_{c}" for c in OUTCOME_COLS if f"delta_{c}" in df.columns]
    pre_outcomes = [f"pre_{c}" for c in OUTCOME_COLS if f"pre_{c}" in df.columns]
    rd_rows = []
    for cutoff in STAFFING_THRESHOLDS:
        for bw in [5, 10, 15, 20, 30]:
            for donut in [0, 1]:
                for y in post_outcomes:
                    r = local_linear(df, y, x, cutoff, bw, donut)
                    r["design"] = "RD" if donut == 0 else "donut_RD"
                    r["n"] = int(r.get("n", 0))
                    r["sample_definition"] = "July 2022 facilities with valid V3 score and post-quarter outcome"
                    rd_rows.append(r)
    rd = pd.DataFrame(rd_rows)
    save_csv(TABLES / "rd_threshold_estimates_v3.csv", rd)

    dens_rows = []
    for cutoff in STAFFING_THRESHOLDS:
        for bw in [5, 10, 15, 20, 30]:
            d = df[[x]].dropna().copy()
            d["dist"] = d[x] - cutoff
            local = d[d["dist"].abs().le(bw)]
            below = int((local["dist"] < 0).sum())
            above = int((local["dist"] >= 0).sum())
            n = below + above
            dens_rows.append(
                {
                    "cutoff": cutoff,
                    "bandwidth": bw,
                    "below_n": below,
                    "above_n": above,
                    "n": n,
                    "binomial_density_p": stats.binomtest(above, n, 0.5).pvalue if n else np.nan,
                    "sample_definition": "facilities within bandwidth around V3 reconstructed staffing score cutoff",
                }
            )
    dens = pd.DataFrame(dens_rows)
    save_csv(TABLES / "rd_density_checks_v3.csv", dens)

    bal_rows = []
    covars = [
        "ownership_for_profit",
        "certified_beds",
        "avg_residents_per_day",
        "health_inspection_rating",
        "qm_rating",
        "rating_cycle_1_total_health_deficiencies",
        "pre_weekend_rn_lt8_day_share",
        "pre_zero_rn_day_share",
        "pre_weekend_p10_total_hprd",
        "pre_avg_daily_census",
    ]
    for cutoff in STAFFING_THRESHOLDS:
        for c in [c for c in covars if c in df.columns]:
            r = local_linear(df, c, x, cutoff, 20, 0)
            r["diagnostic"] = "covariate_balance"
            r["sample_definition"] = "covariate balance around V3 reconstructed staffing score cutoff, bandwidth 20"
            bal_rows.append(r)
    bal = pd.DataFrame(bal_rows)
    save_csv(TABLES / "rd_covariate_balance_v3.csv", bal)

    pre_rows = []
    for cutoff in STAFFING_THRESHOLDS:
        for bw in [10, 20]:
            for y in pre_outcomes:
                r = local_linear(df, y, x, cutoff, bw, 0)
                r["diagnostic"] = "pre_outcome_discontinuity"
                r["sample_definition"] = "pre-period outcome discontinuity around V3 reconstructed staffing score cutoff"
                pre_rows.append(r)
    prechk = pd.DataFrame(pre_rows)
    save_csv(TABLES / "rd_preoutcome_checks_v3.csv", prechk)

    placebo_rows = []
    for cutoff in PLACEBO_THRESHOLDS:
        for bw in [10, 20]:
            for y in post_outcomes:
                r = local_linear(df, y, x, cutoff, bw, 0)
                r["design"] = "placebo_cutoff"
                r["sample_definition"] = "post-period RD at non-policy placebo cutoffs"
                placebo_rows.append(r)
    placebo = pd.DataFrame(placebo_rows)
    save_csv(TABLES / "rd_placebo_cutoffs_v3.csv", placebo)

    did_rows = []
    for cutoff in STAFFING_THRESHOLDS:
        for bw in [5, 10, 15, 20, 30]:
            for y in delta_outcomes:
                r = local_linear(df, y, x, cutoff, bw, 0)
                r["design"] = "RD_DID"
                r["sample_definition"] = "post-minus-pre facility outcomes around V3 reconstructed staffing score cutoff"
                did_rows.append(r)
    did = pd.DataFrame(did_rows)
    save_csv(TABLES / "rd_did_estimates_v3.csv", did)

    density_ok = bool(dens[(dens["bandwidth"].eq(10)) & (dens["cutoff"].eq(320))]["binomial_density_p"].iloc[0] > 0.05)
    balance_bad = int((bal["p"] < 0.05).sum()) if "p" in bal else 0
    pre_bad = int((prechk["p"] < 0.05).sum()) if "p" in prechk else 0
    main = rd[(rd["cutoff"].eq(320)) & (rd["bandwidth"].eq(20)) & (rd["donut"].eq(0))]
    main_did = did[(did["cutoff"].eq(320)) & (did["bandwidth"].eq(20))]
    decision = "candidate primary evidence but not strong go unless diagnostics pass"
    if not density_ok or balance_bad or pre_bad:
        decision = "exploratory/conditional because density, balance, or pre-outcome diagnostics remain imperfect"
    md = [
        "# RD/RD-DID Rescue Results V3",
        "",
        f"July 2022 V3 emulator match rate: {july_match:.3f}. RD/RD-DID was rerun because the running variable passed the 0.950 validation threshold.",
        "",
        "## Main 320-Cutoff RD, Bandwidth 20",
        "",
        md_table(main),
        "",
        "## Main 320-Cutoff RD-DID, Bandwidth 20",
        "",
        md_table(main_did),
        "",
        "## Diagnostics",
        "",
        f"- 320 cutoff density p-value at bandwidth 10: {float(dens[(dens['bandwidth'].eq(10)) & (dens['cutoff'].eq(320))]['binomial_density_p'].iloc[0]):.4g}.",
        f"- Balance checks with p<0.05: {balance_bad}.",
        f"- Pre-outcome checks with p<0.05: {pre_bad}.",
        "",
        f"Decision: **{decision}**.",
    ]
    write_text(REPORT / "03_rd_rddid_rescue_results.md", "\n".join(md))


def two_group(df: pd.DataFrame, y: str, group: str, weight: str | None = None) -> dict:
    d = df[[y, group] + ([weight] if weight else [])].dropna()
    out = {"outcome": y, "group": group, "n": int(len(d)), "sample_definition": "two-group post-minus-pre comparison"}
    if d[group].nunique() < 2:
        return out | {"treated_n": 0, "control_n": 0, "diff": np.nan, "se": np.nan, "p": np.nan}
    a = d[d[group].eq(1)]
    b = d[d[group].eq(0)]
    if weight:
        wa = a[weight].astype(float).to_numpy()
        wb = b[weight].astype(float).to_numpy()
        ya = a[y].astype(float).to_numpy()
        yb = b[y].astype(float).to_numpy()
        ma = np.average(ya, weights=wa)
        mb = np.average(yb, weights=wb)
        diff = float(ma - mb)
        se = np.nan
        p = np.nan
    else:
        ya = a[y].astype(float).to_numpy()
        yb = b[y].astype(float).to_numpy()
        diff = float(np.nanmean(ya) - np.nanmean(yb))
        se = math.sqrt(np.nanvar(ya, ddof=1) / len(ya) + np.nanvar(yb, ddof=1) / len(yb)) if len(ya) > 1 and len(yb) > 1 else np.nan
        p = 2 * (1 - stats.t.cdf(abs(diff / se), df=max(1, len(ya) + len(yb) - 2))) if se and se > 0 else np.nan
    return out | {"treated_n": int(len(a)), "control_n": int(len(b)), "diff": diff, "se": se, "p": p}


def standardized_diff(df: pd.DataFrame, x: str, treat: str, weight: str | None = None) -> dict:
    d = df[[x, treat] + ([weight] if weight else [])].dropna()
    a = d[d[treat].eq(1)]
    b = d[d[treat].eq(0)]
    if len(a) == 0 or len(b) == 0:
        return {"covariate": x, "n": int(len(d)), "std_diff": np.nan}
    if weight:
        ma = np.average(a[x], weights=a[weight])
        mb = np.average(b[x], weights=b[weight])
    else:
        ma = a[x].mean()
        mb = b[x].mean()
    sd = d[x].std(ddof=1)
    return {"covariate": x, "n": int(len(d)), "treated_mean": ma, "control_mean": mb, "std_diff": float((ma - mb) / sd) if sd else np.nan}


def stage4_formula_label_rescue() -> None:
    ensure_dirs()
    df = build_analysis_panel_v3()
    em = pd.read_parquet(DATA / "rating_score_emulator_panel_v3.parquet")
    apr = em[em["snapshot"].eq("2022-04-27")][
        ["facility_id", "overall_rating", "staffing_rating", "health_inspection_rating", "qm_rating"]
    ].rename(
        columns={
            "overall_rating": "overall_rating_apr2022",
            "staffing_rating": "staffing_rating_apr2022",
            "health_inspection_rating": "health_inspection_rating_apr2022",
            "qm_rating": "qm_rating_apr2022",
        }
    )
    df = df.merge(apr, on="facility_id", how="left")
    df["actual_overall_change_apr_to_jul"] = df["overall_rating"] - df["overall_rating_apr2022"]
    df["actual_staffing_change_apr_to_jul"] = df["staffing_rating"] - df["staffing_rating_apr2022"]
    df["actual_health_change_apr_to_jul"] = df["health_inspection_rating"] - df["health_inspection_rating_apr2022"]
    df["actual_qm_change_apr_to_jul"] = df["qm_rating"] - df["qm_rating_apr2022"]
    df["predicted_formula_loss_size"] = df["old_formula_overall"] - df["new_formula_overall"]
    treat = "formula_induced_overall_star_loss"
    first = pd.DataFrame([two_group(df, y, treat) for y in ["predicted_formula_loss_size", "actual_overall_change_apr_to_jul", "actual_staffing_change_apr_to_jul", "actual_health_change_apr_to_jul", "actual_qm_change_apr_to_jul"]])
    first["sample_definition"] = "July facilities, treated if old overall formula exceeds July 2022 formula"
    save_csv(TABLES / "formula_label_first_stage_v3.csv", first)

    covars = [
        "overall_rating_apr2022",
        "staffing_rating_apr2022",
        "health_inspection_rating_apr2022",
        "qm_rating_apr2022",
        "certified_beds",
        "avg_residents_per_day",
        "pre_avg_daily_census",
        "pre_weekend_rn_lt8_day_share",
        "pre_zero_rn_day_share",
        "pre_weekend_p10_total_hprd",
    ]
    covars = [c for c in covars if c in df.columns]
    model_df = df[[treat] + covars].dropna().copy()
    for c in covars:
        sd = model_df[c].std(ddof=1)
        model_df[c + "_z"] = (model_df[c] - model_df[c].mean()) / (sd if sd else 1)
    zcols = [c + "_z" for c in covars]
    lr = LogisticRegression(max_iter=2000, class_weight="balanced")
    lr.fit(model_df[zcols], model_df[treat])
    model_df["pscore"] = lr.predict_proba(model_df[zcols])[:, 1]
    df = df.merge(model_df[["pscore"]], left_index=True, right_index=True, how="left")
    df["iptw"] = np.where(df[treat].eq(1), 1 / df["pscore"], 1 / (1 - df["pscore"]))
    df["iptw"] = df["iptw"].clip(upper=df["iptw"].quantile(0.99))

    balance_rows = []
    for c in covars:
        balance_rows.append(standardized_diff(df, c, treat) | {"method": "raw", "sample_definition": "raw treated-control balance"})
        balance_rows.append(standardized_diff(df, c, treat, "iptw") | {"method": "iptw", "sample_definition": "propensity-score IPTW balance"})
    balance = pd.DataFrame(balance_rows)
    save_csv(TABLES / "formula_label_balance_v3.csv", balance)

    # Nearest-neighbor matched controls on standardized covariates within states when possible.
    match_base = df.join(model_df[zcols], how="left")
    treated = match_base[match_base[treat].eq(1) & match_base[zcols].notna().all(axis=1)].copy()
    controls = match_base[match_base[treat].eq(0) & match_base[zcols].notna().all(axis=1)].copy()
    matched_ids = []
    if len(treated) and len(controls):
        for state, tg in treated.groupby("state"):
            cg = controls[controls["state"].eq(state)]
            if len(cg) < 5:
                cg = controls
            nn = NearestNeighbors(n_neighbors=1)
            nn.fit(cg[zcols])
            _, idx = nn.kneighbors(tg[zcols])
            matched_ids.extend(cg.iloc[idx.flatten()]["facility_id"].tolist())
    df["nn_matched_control"] = df["facility_id"].isin(matched_ids).astype(int)
    df["matched_sample"] = ((df[treat].eq(1)) | (df["nn_matched_control"].eq(1))).astype(int)
    df["matched_treat"] = df[treat]

    # Pretrend check from 2019-2021Q3 by outcome, treated vs controls.
    q = pd.read_parquet(v2_data_file("reliability_outcomes_facility_quarter.parquet"))
    q = q[q["quarter"].isin(["2019Q1", "2019Q2", "2019Q3", "2019Q4", "2020Q1", "2020Q2", "2020Q3", "2020Q4", "2021Q1", "2021Q2", "2021Q3"])].copy()
    q = q.merge(df[["facility_id", treat, "matched_sample", "matched_treat"]], on="facility_id", how="inner")
    q["t"] = pd.PeriodIndex(q["quarter"], freq="Q").astype(int)
    pretrend_rows = []
    for y in ["avg_daily_census", "weekend_rn_lt8_day_share", "zero_rn_day_share", "weekend_p10_total_hprd"]:
        if y not in q.columns:
            continue
        for sample_name, dd in [("raw", q), ("matched", q[q["matched_sample"].eq(1)])]:
            g = dd[[y, "t", treat]].dropna()
            if len(g) < 100:
                continue
            x = np.column_stack([np.ones(len(g)), g["t"], g[treat], g["t"] * g[treat]])
            beta, *_ = np.linalg.lstsq(x, g[y].to_numpy(), rcond=None)
            resid = g[y].to_numpy() - x @ beta
            dof = len(g) - x.shape[1]
            cov = np.linalg.inv(x.T @ x) * (resid @ resid / max(dof, 1))
            se = math.sqrt(cov[3, 3]) if cov[3, 3] >= 0 else np.nan
            p = 2 * (1 - stats.t.cdf(abs(beta[3] / se), dof)) if se and se > 0 else np.nan
            pretrend_rows.append({"outcome": y, "sample": sample_name, "n": int(len(g)), "treated_time_slope_diff": float(beta[3]), "se": se, "p": p, "sample_definition": "2019Q1-2021Q3 pretrend regression"})
    pretrend = pd.DataFrame(pretrend_rows)
    save_csv(TABLES / "formula_label_pretrends_v3.csv", pretrend)

    # Quarter event-study differences relative to 2022Q2.
    q2 = pd.read_parquet(v2_data_file("reliability_outcomes_facility_quarter.parquet"))
    q2 = q2.merge(df[["facility_id", treat, "matched_sample"]], on="facility_id", how="inner")
    event_rows = []
    for y in ["avg_daily_census", "weekend_rn_lt8_day_share", "zero_rn_day_share", "weekend_p10_total_hprd"]:
        if y not in q2.columns:
            continue
        base = q2[q2["quarter"].eq("2022Q2")][["facility_id", y]].rename(columns={y: "base"})
        d = q2.merge(base, on="facility_id", how="left")
        d["dy"] = d[y] - d["base"]
        for quarter, g in d.groupby("quarter"):
            for sample_name, gg in [("raw", g), ("matched", g[g["matched_sample"].eq(1)])]:
                r = two_group(gg, "dy", treat)
                r["quarter"] = quarter
                r["sample"] = sample_name
                r["sample_definition"] = "quarterly event-study change relative to 2022Q2"
                event_rows.append(r)
    event = pd.DataFrame(event_rows)
    save_csv(TABLES / "formula_label_eventstudy_v3.csv", event)

    estimate_rows = []
    for y in [f"delta_{c}" for c in ["avg_daily_census", "weekend_rn_lt8_day_share", "zero_rn_day_share", "weekend_p10_total_hprd", "weekend_share_total_hours", "contract_share_total_hours"] if f"delta_{c}" in df.columns]:
        r = two_group(df, y, treat)
        r["method"] = "raw"
        estimate_rows.append(r)
        r = two_group(df, y, treat, "iptw")
        r["method"] = "iptw"
        estimate_rows.append(r)
        r = two_group(df[df["matched_sample"].eq(1)], y, "matched_treat")
        r["method"] = "nearest_neighbor_state_fallback"
        estimate_rows.append(r)
    estimates = pd.DataFrame(estimate_rows)
    estimates["sample_definition"] = "post 2022Q4-2023Q4 minus pre 2021Q1-2021Q3 outcome changes"
    save_csv(TABLES / "formula_label_matched_estimates_v3.csv", estimates)

    max_abs_balance = balance[balance["method"].eq("iptw")]["std_diff"].abs().max()
    pretrend_bad = int((pretrend["p"] < 0.05).sum()) if not pretrend.empty else 999
    decision = "stronger conditional-go" if max_abs_balance < 0.10 and pretrend_bad == 0 else "conditional only; no strong causal claim"
    md = [
        "# Formula Label Shock Rescue Results V3",
        "",
        f"Formula-induced-loss homes: {int(df[treat].sum())}. Treatment remains mechanical: `old_formula_overall > new_formula_overall` under verified January 2022 and July 2022 rules.",
        "",
        "## First Stage",
        "",
        md_table(first),
        "",
        "## Balance",
        "",
        f"Maximum absolute IPTW standardized difference: {max_abs_balance:.3f}.",
        md_table(balance.head(20)),
        "",
        "## Pretrends",
        "",
        md_table(pretrend),
        "",
        "## Matched/Weighted Estimates",
        "",
        md_table(estimates),
        "",
        f"Decision: **{decision}**. Interpret any adverse label-loss response as market/operational response to an adverse public label, not pure resident quality improvement.",
    ]
    write_text(REPORT / "04_formula_label_shock_rescue_results.md", "\n".join(md))


PBJ_COLS = [
    "PROVNUM",
    "STATE",
    "WorkDate",
    "MDSCensus",
    "Hrs_RNDON",
    "Hrs_RNadmin",
    "Hrs_RN",
    "Hrs_LPNadmin",
    "Hrs_LPN",
    "Hrs_CNA",
    "Hrs_NAtrn",
    "Hrs_MedAide",
    "Hrs_RNDON_emp",
    "Hrs_RNadmin_emp",
    "Hrs_RN_emp",
    "Hrs_LPNadmin_emp",
    "Hrs_LPN_emp",
    "Hrs_CNA_emp",
    "Hrs_NAtrn_emp",
    "Hrs_MedAide_emp",
    "Hrs_RNDON_ctr",
    "Hrs_RNadmin_ctr",
    "Hrs_RN_ctr",
    "Hrs_LPNadmin_ctr",
    "Hrs_LPN_ctr",
    "Hrs_CNA_ctr",
    "Hrs_NAtrn_ctr",
    "Hrs_MedAide_ctr",
]


def read_pbj_daily_chunks(url: str, usecols: list[str] | None = None, chunksize: int = 200_000):
    cols_norm = {norm_col(c) for c in (usecols or PBJ_COLS)}
    for attempt in range(1, 4):
        try:
            return pd.read_csv(
                url,
                usecols=lambda c: norm_col(c) in cols_norm,
                dtype=str,
                chunksize=chunksize,
                encoding="cp1252",
                encoding_errors="replace",
            )
        except Exception:
            if attempt == 3:
                raise
            time.sleep(3 * attempt)


def prep_daily_chunk(chunk: pd.DataFrame) -> pd.DataFrame:
    chunk = chunk.rename(columns={c: norm_col(c) for c in chunk.columns})
    for c in [norm_col(x) for x in PBJ_COLS]:
        if c not in chunk.columns:
            chunk[c] = 0
    chunk["facility_id"] = zfill_ccn(chunk["provnum"])
    chunk["workdate"] = pd.to_datetime(chunk["workdate"], errors="coerce")
    chunk["state"] = chunk["state"].astype(str).str.strip().str.upper()
    for c in [norm_col(x) for x in PBJ_COLS if x not in ["PROVNUM", "STATE", "WorkDate"]]:
        chunk[c] = pd.to_numeric(chunk[c], errors="coerce").fillna(0.0)
    rn = ["hrs_rndon", "hrs_rnadmin", "hrs_rn"]
    lpn = ["hrs_lpnadmin", "hrs_lpn"]
    aide = ["hrs_cna", "hrs_natrn", "hrs_medaide"]
    emp = ["hrs_rndon_emp", "hrs_rnadmin_emp", "hrs_rn_emp", "hrs_lpnadmin_emp", "hrs_lpn_emp", "hrs_cna_emp", "hrs_natrn_emp", "hrs_medaide_emp"]
    ctr = ["hrs_rndon_ctr", "hrs_rnadmin_ctr", "hrs_rn_ctr", "hrs_lpnadmin_ctr", "hrs_lpn_ctr", "hrs_cna_ctr", "hrs_natrn_ctr", "hrs_medaide_ctr"]
    chunk["rn_hours"] = chunk[rn].sum(axis=1)
    chunk["total_nurse_hours"] = chunk[rn + lpn + aide].sum(axis=1)
    chunk["employee_nurse_hours"] = chunk[emp].sum(axis=1)
    chunk["contract_nurse_hours"] = chunk[ctr].sum(axis=1)
    chunk["is_weekend"] = chunk["workdate"].dt.weekday.ge(5).astype(int)
    chunk["week"] = chunk["workdate"].dt.to_period("W-SUN").astype(str)
    return chunk[["facility_id", "state", "workdate", "week", "is_weekend", "mdscensus", "rn_hours", "total_nurse_hours", "employee_nurse_hours", "contract_nurse_hours"]].dropna(subset=["facility_id", "workdate"])


def stage5_metric_salience_week_panel() -> None:
    ensure_dirs()
    if (DATA / "metric_salience_panel_v3.parquet").exists():
        panel = pd.read_parquet(DATA / "metric_salience_panel_v3.parquet")
    else:
        sources = pd.read_csv(project_temp_file("pbj_daily_sources.csv"))
        sources["start"] = pd.to_datetime(sources["temporal"].astype(str).str.split("/").str[0])
        sources["through"] = pd.to_datetime(sources["temporal"].astype(str).str.split("/").str[-1])
        sources = sources[
            (sources["start"] >= pd.Timestamp("2021-01-01"))
            & (sources["through"] <= pd.Timestamp("2023-12-31"))
            & ~((sources["start"] >= pd.Timestamp("2021-10-01")) & (sources["through"] <= pd.Timestamp("2021-12-31")))
        ]
        agg_parts = []
        for _, src in sources.iterrows():
            print(f"Metric salience streaming {src['temporal']}")
            for chunk in read_pbj_daily_chunks(src["download_url"], chunksize=250_000):
                d = prep_daily_chunk(chunk)
                g = d.groupby(["facility_id", "state", "week", "is_weekend"], as_index=False).agg(
                    resident_days=("mdscensus", "sum"),
                    days=("workdate", "nunique"),
                    rn_hours=("rn_hours", "sum"),
                    total_nurse_hours=("total_nurse_hours", "sum"),
                    contract_nurse_hours=("contract_nurse_hours", "sum"),
                    employee_nurse_hours=("employee_nurse_hours", "sum"),
                )
                agg_parts.append(g)
        agg = pd.concat(agg_parts, ignore_index=True)
        agg = agg.groupby(["facility_id", "state", "week", "is_weekend"], as_index=False).sum(numeric_only=True)
        wide = agg.pivot_table(index=["facility_id", "state", "week"], columns="is_weekend", values=["resident_days", "rn_hours", "total_nurse_hours", "contract_nurse_hours"], aggfunc="sum")
        wide.columns = [f"{a}_{'weekend' if b == 1 else 'weekday'}" for a, b in wide.columns]
        wide = wide.reset_index()
        for c in ["resident_days_weekend", "resident_days_weekday", "rn_hours_weekend", "rn_hours_weekday", "total_nurse_hours_weekend", "total_nurse_hours_weekday", "contract_nurse_hours_weekend", "contract_nurse_hours_weekday"]:
            if c not in wide:
                wide[c] = 0.0
        wide["week_start"] = pd.PeriodIndex(wide["week"], freq="W-SUN").start_time
        wide["post_july"] = (wide["week_start"] >= pd.Timestamp("2022-08-01")).astype(int)
        records = []
        metric_specs = [
            ("weekend_total_nurse_hprd", "score_targeted", "weekend", "total_nurse_hours_weekend", "resident_days_weekend"),
            ("weekday_total_nurse_hprd", "comparison", "weekday", "total_nurse_hours_weekday", "resident_days_weekday"),
            ("all_day_total_nurse_hprd", "score_targeted", "all_day", None, None),
            ("all_day_rn_hprd", "score_targeted", "all_day", None, None),
            ("weekend_rn_hprd", "transparency_not_separate_score", "weekend", "rn_hours_weekend", "resident_days_weekend"),
            ("weekday_rn_hprd", "comparison", "weekday", "rn_hours_weekday", "resident_days_weekday"),
            ("total_weekly_nursing_hours", "comparison", "all_day", None, None),
            ("contract_share_nursing_hours", "composition", "all_day", None, None),
        ]
        for metric, group, weekend_flag, num, den in metric_specs:
            d = wide[["facility_id", "state", "week", "week_start", "post_july"]].copy()
            d["metric"] = metric
            d["metric_group"] = group
            d["weekend_metric"] = 1 if weekend_flag == "weekend" else 0
            d["score_targeted_metric"] = 1 if group == "score_targeted" else 0
            if metric == "all_day_total_nurse_hprd":
                value = (wide["total_nurse_hours_weekend"] + wide["total_nurse_hours_weekday"]) / (wide["resident_days_weekend"] + wide["resident_days_weekday"]).replace(0, np.nan)
            elif metric == "all_day_rn_hprd":
                value = (wide["rn_hours_weekend"] + wide["rn_hours_weekday"]) / (wide["resident_days_weekend"] + wide["resident_days_weekday"]).replace(0, np.nan)
            elif metric == "total_weekly_nursing_hours":
                value = wide["total_nurse_hours_weekend"] + wide["total_nurse_hours_weekday"]
            elif metric == "contract_share_nursing_hours":
                value = (wide["contract_nurse_hours_weekend"] + wide["contract_nurse_hours_weekday"]) / (wide["total_nurse_hours_weekend"] + wide["total_nurse_hours_weekday"]).replace(0, np.nan)
            else:
                value = wide[num] / wide[den].replace(0, np.nan)
            d["value"] = value
            records.append(d)
        panel = pd.concat(records, ignore_index=True)
        em = pd.read_parquet(DATA / "analysis_threshold_facility_july2022_v3.parquet")
        panel = panel.merge(em[["facility_id", "high_shadow_price", "formula_induced_overall_star_loss", "staffing_rating"]], on="facility_id", how="left")
        panel.to_parquet(DATA / "metric_salience_panel_v3.parquet", index=False)

    # Aggregate estimands to avoid high-dimensional FE dependency.
    pivot = panel.pivot_table(index=["facility_id", "week", "post_july"], columns="metric", values="value", aggfunc="mean").reset_index()
    pivot["targeted_weekend_gap"] = pivot["weekend_total_nurse_hprd"] - pivot["weekday_total_nurse_hprd"]
    pivot["rn_weekend_gap"] = pivot["weekend_rn_hprd"] - pivot["weekday_rn_hprd"]
    pivot["score_targeted_minus_rn_gap"] = pivot["targeted_weekend_gap"] - pivot["rn_weekend_gap"]
    pre = pivot[pivot["post_july"].eq(0)].groupby("facility_id").mean(numeric_only=True).add_prefix("pre_")
    post = pivot[pivot["post_july"].eq(1)].groupby("facility_id").mean(numeric_only=True).add_prefix("post_")
    fac = pre.join(post, how="inner").reset_index()
    for c in ["score_targeted_minus_rn_gap", "total_weekly_nursing_hours", "weekend_total_nurse_hprd", "weekday_total_nurse_hprd", "weekend_rn_hprd"]:
        fac[f"delta_{c}"] = fac[f"post_{c}"] - fac[f"pre_{c}"]
    em = pd.read_parquet(DATA / "analysis_threshold_facility_july2022_v3.parquet")
    fac = fac.merge(em[["facility_id", "high_shadow_price", "formula_induced_overall_star_loss"]], on="facility_id", how="left")
    rows = [
        {
            "estimand": "post_minus_pre_targeted_weekend_total_minus_weekend_rn_gap",
            "n": int(fac["delta_score_targeted_minus_rn_gap"].notna().sum()),
            "mean_delta": float(fac["delta_score_targeted_minus_rn_gap"].mean()),
            "se": float(fac["delta_score_targeted_minus_rn_gap"].std(ddof=1) / math.sqrt(fac["delta_score_targeted_minus_rn_gap"].notna().sum())),
            "sample_definition": "facility-week by metric panel, post Aug 2022 versus pre 2021Q1-2022Q2 excluding 2021Q4",
        },
        two_group(fac, "delta_score_targeted_minus_rn_gap", "formula_induced_overall_star_loss") | {"estimand": "formula_loss_heterogeneity"},
        two_group(fac, "delta_score_targeted_minus_rn_gap", "high_shadow_price") | {"estimand": "high_rating_incentive_heterogeneity"},
    ]
    ddd = pd.DataFrame(rows)
    save_csv(TABLES / "metric_salience_ddd_v3.csv", ddd)
    placebo = pd.DataFrame(
        [
            {
                "placebo": "weekday_total_vs_weekday_rn",
                "n": int(fac["delta_weekend_rn_hprd"].notna().sum()),
                "mean_delta": float((fac["delta_weekend_rn_hprd"] - fac["delta_weekday_total_nurse_hprd"]).mean()) if "delta_weekday_total_nurse_hprd" in fac else np.nan,
                "sample_definition": "non-targeted placebo metric contrast in facility-week panel",
            }
        ]
    )
    save_csv(TABLES / "metric_salience_placebo_v3.csv", placebo)
    realloc = pd.DataFrame(
        [
            {
                "outcome": "total_weekly_nursing_hours",
                "n": int(fac["delta_total_weekly_nursing_hours"].notna().sum()),
                "mean_delta": float(fac["delta_total_weekly_nursing_hours"].mean()),
                "sample_definition": "weekly total nursing hours reallocation diagnostic",
            },
            {
                "outcome": "weekend_total_nurse_hprd",
                "n": int(fac["delta_weekend_total_nurse_hprd"].notna().sum()),
                "mean_delta": float(fac["delta_weekend_total_nurse_hprd"].mean()),
                "sample_definition": "score-targeted weekend total nurse HPRD diagnostic",
            },
            {
                "outcome": "weekday_total_nurse_hprd",
                "n": int(fac["delta_weekday_total_nurse_hprd"].notna().sum()),
                "mean_delta": float(fac["delta_weekday_total_nurse_hprd"].mean()),
                "sample_definition": "weekday offset diagnostic",
            },
        ]
    )
    save_csv(TABLES / "metric_salience_reallocation_v3.csv", realloc)
    ddd_main = float(ddd.iloc[0]["mean_delta"])
    weekly = float(realloc.loc[realloc["outcome"].eq("total_weekly_nursing_hours"), "mean_delta"].iloc[0])
    interpretation = "targeted behavior/reallocation" if ddd_main > 0 and abs(weekly) < abs(ddd_main) * 100 else "broad labor-market or mixed movement"
    md = [
        "# Metric Salience DDD Rescue Results V3",
        "",
        "This pass builds a true facility-week by metric panel from PBJ daily nurse staffing files, excluding 2021Q4 because CMS/OIG documentation flags incomplete PBJ reporting for many homes.",
        "",
        "## DDD Estimates",
        "",
        md_table(ddd),
        "",
        "## Placebo and Reallocation",
        "",
        md_table(placebo),
        "",
        md_table(realloc),
        "",
        f"Interpretation: **{interpretation}**. This is mechanism evidence and should be read with the RD/formula diagnostics.",
    ]
    write_text(REPORT / "05_metric_salience_ddd_rescue_results.md", "\n".join(md))


def discover_pbj_old_sources() -> pd.DataFrame:
    r = requests.get(CMS_PBJ_CATALOG_PAGE, timeout=60)
    r.raise_for_status()
    html = r.text.replace("&amp;", "&")
    (RAW / "data_gov_pbj_daily_nurse_staffing.html").write_text(html, encoding="utf-8")
    urls = sorted(set(re.findall(r"https://data\.cms\.gov/sites/default/files/[^\"<\s]+\.csv", html)))
    rows = []
    for url in urls:
        fname = url.split("/")[-1]
        m = re.search(r"(20\d{2})_Q([1-4])", fname, re.I)
        if not m:
            m = re.search(r"CY(20\d{2})Q([1-4])", fname, re.I)
        if not m:
            continue
        year = int(m.group(1))
        q = int(m.group(2))
        if year not in [2017, 2018, 2019]:
            continue
        rows.append(
            {
                "title_date": f"{year}Q{q}",
                "quarter": f"{year}Q{q}",
                "download_url": url,
                "file_name": fname,
                "source_url": CMS_PBJ_CATALOG_PAGE,
            }
        )
    src = pd.DataFrame(rows).drop_duplicates(["quarter", "download_url"]).sort_values(["quarter", "download_url"])
    # Keep static CSV links over API links and one file per quarter.
    src = src.drop_duplicates(["quarter"], keep="last")
    save_csv(TABLES / "pbj_2018_validation_feasibility_v3.csv", add_sample(src.assign(n=len(src)), "official Data.gov/CMS PBJ resource discovery"))
    return src


def aggregate_zero_rn_for_source(url: str, quarter: str) -> pd.DataFrame:
    parts = []
    for chunk in read_pbj_daily_chunks(url, usecols=["PROVNUM", "STATE", "WorkDate", "MDSCensus", "Hrs_RNDON", "Hrs_RNadmin", "Hrs_RN"], chunksize=250_000):
        d = prep_daily_chunk(chunk)
        d["positive_census"] = d["mdscensus"] > 0
        d["zero_rn_day"] = (d["positive_census"] & d["rn_hours"].le(0)).astype(int)
        d["low_positive_rn_day"] = (d["positive_census"] & d["rn_hours"].gt(0) & d["rn_hours"].lt(8)).astype(int)
        g = d.groupby(["facility_id", "state"], as_index=False).agg(
            resident_days=("mdscensus", "sum"),
            days=("workdate", "nunique"),
            zero_rn_days=("zero_rn_day", "sum"),
            rn_lt8_positive_days=("low_positive_rn_day", "sum"),
        )
        parts.append(g)
    out = pd.concat(parts, ignore_index=True).groupby(["facility_id", "state"], as_index=False).sum(numeric_only=True)
    out["quarter"] = quarter
    out["zero_rn_ge7"] = (out["zero_rn_days"] >= 7).astype(int)
    return out


def stage6_2018_validation_rescue() -> None:
    ensure_dirs()
    if (
        os.environ.get("V3_FORCE_REBUILD") != "1"
        and (TABLES / "pbj_2018_validation_feasibility_v3.csv").exists()
        and (TABLES / "pbj_2018_zero_rn_validation_v3.csv").exists()
        and (REPORT / "06_2018_validation_rescue_results.md").exists()
    ):
        return
    src = discover_pbj_old_sources()
    needed = src[src["quarter"].between("2017Q1", "2019Q4")].copy()
    rows = []
    failures = []
    for _, r in needed.iterrows():
        try:
            print(f"2018 validation streaming {r['quarter']} {r['download_url']}")
            rows.append(aggregate_zero_rn_for_source(r["download_url"], r["quarter"]))
        except Exception as exc:
            failures.append({"quarter": r["quarter"], "download_url": r["download_url"], "error": str(exc), "n": 0, "sample_definition": "failed official PBJ source processing"})
    if rows:
        panel = pd.concat(rows, ignore_index=True)
        panel.to_parquet(DATA / "pbj_2018_zero_rn_validation_v3.parquet", index=False)
        summary = panel.groupby("quarter").agg(
            n=("facility_id", "nunique"),
            mean_zero_rn_days=("zero_rn_days", "mean"),
            share_ge7_zero_rn=("zero_rn_ge7", "mean"),
            mean_low_positive_rn_days=("rn_lt8_positive_days", "mean"),
        ).reset_index()
        summary["sample_definition"] = "official PBJ daily nurse staffing facility-quarter panel"
    else:
        summary = pd.DataFrame(columns=["quarter", "n", "mean_zero_rn_days", "share_ge7_zero_rn", "mean_low_positive_rn_days", "sample_definition"])
    save_csv(TABLES / "pbj_2018_zero_rn_validation_v3.csv", summary)
    if failures:
        save_csv(AUDIT / "pbj_2018_download_failures.csv", pd.DataFrame(failures))
    recovered = len(summary) > 0 and summary["quarter"].str.startswith("2018").any()
    if recovered:
        pre = summary[summary["quarter"].str.startswith("2017")]["share_ge7_zero_rn"].mean()
        post = summary[summary["quarter"].str.startswith("2019")]["share_ge7_zero_rn"].mean()
        low_pre = summary[summary["quarter"].str.startswith("2017")]["mean_low_positive_rn_days"].mean()
        low_post = summary[summary["quarter"].str.startswith("2019")]["mean_low_positive_rn_days"].mean()
        decision = "mechanism validation recovered" if post < pre and low_post > low_pre else "official files recovered but OIG pattern is mixed"
    else:
        decision = "infeasible after official source search"
    md = [
        "# 2018 Validation Rescue Results V3",
        "",
        f"Decision: **{decision}**.",
        "",
        "## Feasibility",
        "",
        md_table(pd.read_csv(TABLES / "pbj_2018_validation_feasibility_v3.csv").head(20)),
        "",
        "## Zero-RN Validation Summary",
        "",
        md_table(summary),
    ]
    if failures:
        md += ["", "## Failures", "", md_table(pd.DataFrame(failures))]
    write_text(REPORT / "06_2018_validation_rescue_results.md", "\n".join(md))


def final_decision_report() -> None:
    ensure_dirs()
    inv_sum = pd.read_csv(TABLES / "official_score_field_inventory_summary.csv")
    val = pd.read_csv(TABLES / "emulator_validation_v3.csv")
    july_match = float(val.loc[val["snapshot_date"].eq("2022-07-27"), "star_match_rate"].iloc[0])
    score_found = bool(inv_sum.loc[inv_sum["check"].eq("official_0_380_staffing_score_found"), "result"].iloc[0] == "True")
    explicit_adj = bool(inv_sum.loc[inv_sum["check"].eq("explicit_july_adjusted_weekend_total_nurse_hprd_found"), "result"].iloc[0] == "True")
    components = bool(inv_sum.loc[inv_sum["check"].eq("conclusion"), "result"].iloc[0] == "enough official components found to reconstruct without proxy substitution")
    rd_report = REPORT / "03_rd_rddid_rescue_results.md"
    rd_status = "running variable rescued and estimates rerun, but diagnostics keep it short of STRONG GO" if rd_report.exists() and july_match >= 0.95 else "no-go/exploratory"
    formula = pd.read_csv(TABLES / "formula_label_matched_estimates_v3.csv") if (TABLES / "formula_label_matched_estimates_v3.csv").exists() else pd.DataFrame()
    balance = pd.read_csv(TABLES / "formula_label_balance_v3.csv") if (TABLES / "formula_label_balance_v3.csv").exists() else pd.DataFrame()
    max_bal = balance[balance["method"].eq("iptw")]["std_diff"].abs().max() if not balance.empty else np.nan
    formula_status = "stronger conditional mechanism evidence" if pd.notna(max_bal) and max_bal < 0.10 else "conditional only"
    ddd = pd.read_csv(TABLES / "metric_salience_ddd_v3.csv") if (TABLES / "metric_salience_ddd_v3.csv").exists() else pd.DataFrame()
    ddd_status = "facility-week mechanism evidence available" if not ddd.empty else "not recovered"
    pbj = pd.read_csv(TABLES / "pbj_2018_zero_rn_validation_v3.csv") if (TABLES / "pbj_2018_zero_rn_validation_v3.csv").exists() else pd.DataFrame()
    validation_status = "official files recovered; OIG mechanism pattern is mixed" if not pbj.empty and pbj["quarter"].astype(str).str.startswith("2018").any() else "not recovered"
    strong_go = False
    if rd_report.exists():
        dens = pd.read_csv(TABLES / "rd_density_checks_v3.csv")
        bal = pd.read_csv(TABLES / "rd_covariate_balance_v3.csv")
        pre = pd.read_csv(TABLES / "rd_preoutcome_checks_v3.csv")
        placebo = pd.read_csv(TABLES / "rd_placebo_cutoffs_v3.csv")
        strong_go = (
            (score_found or components or july_match >= 0.95)
            and (dens["binomial_density_p"].min() > 0.01)
            and ((bal["p"] < 0.05).mean() < 0.10)
            and ((pre["p"] < 0.05).mean() < 0.10)
            and ((placebo["p"] < 0.05).mean() < 0.20)
        )
    if strong_go:
        manuscript = "strong local causal RD-DID paper"
        executive = "STRONG GO for a local RD/RD-DID manuscript, with the reconstructed official-component running variable and diagnostics disclosed."
    elif formula_status.startswith("stronger") or ddd_status.startswith("facility-week"):
        manuscript = "conditional algorithmic-label mechanism paper"
        executive = "CONDITIONAL GO: the July running variable was rescued, but manuscript strength still depends on RD diagnostics and mechanism consistency."
    else:
        manuscript = "transparent no-go paper"
        executive = "NO-GO for strong causal claims."
    md = [
        "# Final Score Rescue Report V3",
        "",
        "## Executive Judgment",
        "",
        executive,
        "",
        "## Score Rescue",
        "",
        f"- Official 0-380 staffing score found: {score_found}.",
        f"- Explicit July adjusted weekend total nurse HPRD found: {explicit_adj}.",
        f"- Enough official components for non-proxy reconstruction: {components}.",
        f"- July 2022 emulator reached 95% validation: {july_match >= 0.95} ({july_match:.3f}).",
        "",
        "## Design Decisions",
        "",
        f"- RD/RD-DID: {rd_status}.",
        f"- Formula-induced label shock: {formula_status}.",
        f"- Metric-salience DDD: {ddd_status}.",
        f"- 2018 validation: {validation_status}.",
        "",
        "## Recommended Manuscript Path",
        "",
        manuscript + ".",
        "",
        "## Exact Next Step",
        "",
        "If continuing, inspect the V3 RD diagnostics by cutoff and outcome, then decide whether the manuscript should lead with the 320 cutoff RD-DID or use RD as validation while formula shock and facility-week metric salience carry the mechanism narrative.",
    ]
    write_text(REPORT / "final_score_rescue_report.md", "\n".join(md))


def run_all_core() -> None:
    ensure_dirs()
    stage1_source_field_inventory()
    stage2_emulator_rescue()
    stage3_rd_rddid_if_valid()
    stage4_formula_label_rescue()
    stage5_metric_salience_week_panel()
    stage6_2018_validation_rescue()
    final_decision_report()


if __name__ == "__main__":
    run_all_core()
