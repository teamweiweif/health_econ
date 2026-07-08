from __future__ import annotations

import csv
import io
import json
import math
import re
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook


THIS_FILE = Path(__file__).resolve()
AUDIT_ROOT = THIS_FILE.parents[1]
PROJECT_ROOT = AUDIT_ROOT.parent
REPO_ROOT = PROJECT_ROOT.parent

DATA = AUDIT_ROOT / "data"
TABLES = AUDIT_ROOT / "result" / "tables"
REPORT = AUDIT_ROOT / "report"
SNAP = AUDIT_ROOT / "temp" / "source_snapshots"
AUDIT_LOGS = AUDIT_ROOT / "temp" / "audit_logs"

DESIGN_TOURNAMENT = PROJECT_ROOT / "design_tournament"
V3_ROOT = DESIGN_TOURNAMENT / "v3_score_rescue"
SIBLING_PROJECT = REPO_ROOT.parent / "nh_staffing" / "nursing_home_staffing_reporting"
SIBLING_PROVIDER_ARCHIVES = SIBLING_PROJECT / "temp" / "raw_downloads" / "provider_archives"
LOCAL_PROVIDER_ARCHIVES = PROJECT_ROOT / "temp" / "raw_downloads" / "provider_archives"

CMS_ARCHIVE_API = "https://data.cms.gov/provider-data/api/1/archive/aggregate/theme/nursing-homes"
PBJ_DAILY_VIEWER = "https://data.cms.gov/data-api/v1/dataset/7e0d53ba-8f02-4c66-98a5-14a1c997c50d/data-viewer"
PBJ_DAILY_DATA = "https://data.cms.gov/data-api/v1/dataset/7e0d53ba-8f02-4c66-98a5-14a1c997c50d/data?size=1"
PBJ_EMP_VIEWER = "https://data.cms.gov/data-api/v1/dataset/d65b8be0-946e-410b-ab06-01829628d5a1/data-viewer"
PBJ_EMP_DATA = "https://data.cms.gov/data-api/v1/dataset/d65b8be0-946e-410b-ab06-01829628d5a1/data?size=1"
PBJ_PUF_DOC = "https://data.cms.gov/sites/default/files/2022-04/dcd6331a-5f30-4f4a-b324-bcac40bada95/PBJ_PUF_Documentation_April_2022.pdf"
PBJ_EMP_DOC = "https://download.cms.gov/pbj/pbj_employeedetailpuf_documentation_january_2022.pdf"
CMS_PBJ_PAGE = "https://www.cms.gov/medicare/quality/nursing-home-improvement/staffing-data-submission"
QSO_2208 = "https://www.cms.gov/files/document/qso-22-08-nh.pdf"
JULY_CARE_COMPARE = "https://www.cms.gov/newsroom/fact-sheets/updates-care-compare-website-july-2022"

HISTORICAL_SNAPSHOT_SPECS = [
    ("January 2022", "2022-01-27"),
    ("April 2022", "2022-04-27"),
    ("July 2022", "2022-07-27"),
    ("October 2022", "2022-10-27"),
    ("January 2023", "2023-01-02"),
]


def ensure_dirs() -> None:
    for path in [DATA, TABLES, REPORT, SNAP, AUDIT_LOGS, SNAP / "provider_archives"]:
        path.mkdir(parents=True, exist_ok=True)


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if columns is None:
        columns = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def fetch_json(url: str, out_path: Path) -> dict[str, Any]:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    last_error: Exception | None = None
    for _ in range(3):
        try:
            response = requests.get(url, timeout=120)
            response.raise_for_status()
            out_path.write_bytes(response.content)
            return response.json()
        except Exception as exc:
            last_error = exc
    if out_path.exists():
        return json.loads(out_path.read_text(encoding="utf-8"))
    raise RuntimeError(f"Could not fetch {url}") from last_error


def fetch_text_json(url: str, out_path: Path) -> dict[str, Any]:
    data = fetch_json(url, out_path)
    return data


def provider_archive_dir() -> Path:
    if SIBLING_PROVIDER_ARCHIVES.exists():
        return SIBLING_PROVIDER_ARCHIVES
    if LOCAL_PROVIDER_ARCHIVES.exists():
        return LOCAL_PROVIDER_ARCHIVES
    raise FileNotFoundError("Provider Data Catalog archives were not found in sibling or repo temp paths.")


def archive_api_records() -> list[dict[str, Any]]:
    api = fetch_json(CMS_ARCHIVE_API, SNAP / "cms_provider_archive_api_nursing_homes.json")
    records = api.get("data", [])
    if not isinstance(records, list):
        raise ValueError("Unexpected Provider Data Catalog archive API structure.")
    return records


def latest_theme_snapshot(records: list[dict[str, Any]]) -> dict[str, Any]:
    snapshots = [r for r in records if r.get("type") == "theme" and r.get("theme") == "nursing-homes"]
    if not snapshots:
        raise ValueError("No nursing-homes theme snapshots returned by CMS archive API.")
    snapshots.sort(key=lambda r: str(r.get("date", "")), reverse=True)
    return snapshots[0]


def archive_record_for(records: list[dict[str, Any]], snapshot_date: str) -> dict[str, Any] | None:
    for rec in records:
        if rec.get("type") == "theme" and rec.get("date") == snapshot_date:
            return rec
    return None


def provider_zip(snapshot_date: str, records: list[dict[str, Any]]) -> tuple[Path, str]:
    local = provider_archive_dir() / f"nursing_homes_{snapshot_date}.zip"
    if local.exists():
        return local, str(local)
    cached = SNAP / "provider_archives" / f"nursing_homes_{snapshot_date}.zip"
    if cached.exists():
        return cached, str(cached)
    rec = archive_record_for(records, snapshot_date)
    if not rec or not rec.get("url"):
        raise FileNotFoundError(f"No local ZIP or CMS archive record found for {snapshot_date}")
    response = requests.get(rec["url"], timeout=300)
    response.raise_for_status()
    cached.write_bytes(response.content)
    return cached, rec["url"]


def provider_info_member(zf: zipfile.ZipFile) -> str:
    members = [m for m in zf.namelist() if "NH_ProviderInfo" in m and m.lower().endswith(".csv")]
    if not members:
        raise FileNotFoundError("NH_ProviderInfo CSV missing from Provider Data Catalog archive.")
    return members[0]


def data_dictionary_member(zf: zipfile.ZipFile) -> str:
    members = [m for m in zf.namelist() if m.endswith("NH_Primary_Data_Dictionary.xlsx")]
    if not members:
        raise FileNotFoundError("NH_Primary_Data_Dictionary.xlsx missing from Provider Data Catalog archive.")
    return members[0]


def read_provider_header(zp: Path) -> tuple[str, list[str]]:
    with zipfile.ZipFile(zp) as zf:
        member = provider_info_member(zf)
        with zf.open(member) as f:
            header = next(csv.reader(io.TextIOWrapper(f, encoding="cp1252", errors="replace")))
    return member, header


def read_provider_info(zp: Path, usecols: list[str]) -> pd.DataFrame:
    with zipfile.ZipFile(zp) as zf:
        member = provider_info_member(zf)
        with zf.open(member) as f:
            return pd.read_csv(f, usecols=[c for c in usecols if c], encoding="cp1252", encoding_errors="replace", low_memory=False)


def read_provider_dictionary(zp: Path) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    with zipfile.ZipFile(zp) as zf:
        try:
            member = data_dictionary_member(zf)
        except FileNotFoundError:
            return out
        wb = load_workbook(io.BytesIO(zf.read(member)), read_only=True, data_only=True)
        for ws in wb.worksheets:
            sheet: dict[str, str] = {}
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1:
                    continue
                label = "" if row[0] is None else str(row[0]).strip()
                desc = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
                if label:
                    sheet[label] = desc
            out[ws.title] = sheet
    return out


def scan_score_fields(zp: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with zipfile.ZipFile(zp) as zf:
        for member in zf.namelist():
            if not member.lower().endswith(".csv"):
                continue
            try:
                with zf.open(member) as f:
                    header = next(csv.reader(io.TextIOWrapper(f, encoding="cp1252", errors="replace")))
            except Exception:
                continue
            for field in header:
                if "score" not in field.lower():
                    continue
                role = "unrelated score fields"
                if re.search(r"q[1-4] measure score|four quarter average score|adjusted score|observed score|expected score", field, re.I):
                    role = "possible score fields"
                rows.append({"file_name": member, "field": field, "role": role})
    return rows


def missingness(df: pd.DataFrame, field: str) -> str:
    if field not in df.columns:
        return "field_absent"
    return f"{float(df[field].isna().mean()):.6f}"


def norm_project_name(field: str) -> str:
    mapping = {
        "Federal Provider Number": "facility_id",
        "Provider Name": "provider_name",
        "Provider State": "state",
        "Ownership Type": "ownership_type",
        "Number of Certified Beds": "certified_beds",
        "Average Number of Residents per Day": "avg_residents_per_day",
        "Reported Total Nurse Staffing Hours per Resident per Day": "reported_total_nurse_hprd",
        "Adjusted Total Nurse Staffing Hours per Resident per Day": "adjusted_total_nurse_hprd",
        "Reported RN Staffing Hours per Resident per Day": "reported_rn_hprd",
        "Adjusted RN Staffing Hours per Resident per Day": "adjusted_rn_hprd",
        "Total number of nurse staff hours per resident per day on the weekend": "reported_weekend_total_nurse_hprd",
        "Registered Nurse hours per resident per day on the weekend": "reported_weekend_rn_hprd",
        "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day": "adjusted_weekend_total_nurse_hprd_official",
        "Total nursing staff turnover": "total_nursing_staff_turnover",
        "Registered Nurse turnover": "rn_turnover",
        "Number of administrators who have left the nursing home": "administrator_departures",
        "Staffing Rating": "staffing_rating",
        "RN Staffing Rating": "rn_staffing_rating",
        "Overall Rating": "overall_rating",
        "Health Inspection Rating": "health_inspection_rating",
        "QM Rating": "qm_rating",
        "Rating Cycle 1 Total Health Score": "rating_cycle_1_total_health_score",
        "Total Weighted Health Survey Score": "total_weighted_health_survey_score",
    }
    return mapping.get(field, field.lower().replace(" ", "_").replace("-", "_"))


def source_ev(*parts: str) -> str:
    return "; ".join([p for p in parts if p])


PBJ_DAILY_LABELS = {
    "PROVNUM": ("Medicare provider number", "", "", "facility-day"),
    "PROVNAME": ("Provider name", "", "", "facility-day"),
    "CITY": ("Provider city", "", "", "facility-day"),
    "STATE": ("Postal abbreviation for state", "", "", "facility-day"),
    "COUNTY_NAME": ("Name of provider county, unique within state", "", "", "facility-day"),
    "COUNTY_FIPS": ("FIPS code for provider county, unique within state", "", "", "facility-day"),
    "CY_Qtr": ("Calendar quarter", "", "", "facility-day"),
    "WorkDate": ("Day for reported hours", "", "YYYYMMDD", "facility-day"),
    "MDScensus": ("Resident census from MDS", "", "resident count", "facility-day"),
    "Hrs_RNDON": ("Total hours for RN Director of Nursing", "5", "hours", "facility-day"),
    "Hrs_RNDON_emp": ("Employee hours for RN Director of Nursing", "5", "hours", "facility-day"),
    "Hrs_RNDON_ctr": ("Contract hours for RN Director of Nursing", "5", "hours", "facility-day"),
    "Hrs_RNadmin": ("Total hours for RN with administrative duties", "6", "hours", "facility-day"),
    "Hrs_RNadmin_emp": ("Employee hours for RN with administrative duties", "6", "hours", "facility-day"),
    "Hrs_RNadmin_ctr": ("Contract hours for RN with administrative duties", "6", "hours", "facility-day"),
    "Hrs_RN": ("Total hours for registered nurse", "7", "hours", "facility-day"),
    "Hrs_RN_emp": ("Employee hours for registered nurse", "7", "hours", "facility-day"),
    "Hrs_RN_ctr": ("Contract hours for registered nurse", "7", "hours", "facility-day"),
    "Hrs_LPNadmin": ("Total hours for LPN with administrative duties", "8", "hours", "facility-day"),
    "Hrs_LPNadmin_emp": ("Employee hours for LPN with administrative duties", "8", "hours", "facility-day"),
    "Hrs_LPNadmin_ctr": ("Contract hours for LPN with administrative duties", "8", "hours", "facility-day"),
    "Hrs_LPN": ("Total hours for licensed practical/vocational nurse", "9", "hours", "facility-day"),
    "Hrs_LPN_emp": ("Employee hours for licensed practical/vocational nurse", "9", "hours", "facility-day"),
    "Hrs_LPN_ctr": ("Contract hours for licensed practical/vocational nurse", "9", "hours", "facility-day"),
    "Hrs_CNA": ("Total hours for certified nurse aide", "10", "hours", "facility-day"),
    "Hrs_CNA_emp": ("Employee hours for certified nurse aide", "10", "hours", "facility-day"),
    "Hrs_CNA_ctr": ("Contract hours for certified nurse aide", "10", "hours", "facility-day"),
    "Hrs_NAtrn": ("Total hours for nurse aide in training", "11", "hours", "facility-day"),
    "Hrs_NAtrn_emp": ("Employee hours for nurse aide in training", "11", "hours", "facility-day"),
    "Hrs_NAtrn_ctr": ("Contract hours for nurse aide in training", "11", "hours", "facility-day"),
    "Hrs_MedAide": ("Total hours for medication aide/technician", "12", "hours", "facility-day"),
    "Hrs_MedAide_emp": ("Employee hours for medication aide/technician", "12", "hours", "facility-day"),
    "Hrs_MedAide_ctr": ("Contract hours for medication aide/technician", "12", "hours", "facility-day"),
}

EMPLOYEE_DETAIL_LABELS = {
    "PROVNUM": ("Federal Provider Number", "facility identifier", "facility-employee-day-jobtype"),
    "STATE": ("Postal abbreviation for state", "state code", "facility-employee-day-jobtype"),
    "CY_Qtr": ("Calendar Quarter", "yyyyQn", "facility-employee-day-jobtype"),
    "WorkDate": ("Work Date", "YYYYMMDD", "facility-employee-day-jobtype"),
    "SYS_EMPLEE_ID": ("System Employee ID", "system-generated employee identifier", "facility-employee-day-jobtype"),
    "EMPLEE_JOB_CD_ID": ("Employee Job Code", "PBJ mandatory reporting job code; includes nursing and non-nursing codes", "facility-employee-day-jobtype"),
    "EMP_CTR": ("Employee Type", "1 = Employee; 2 = Contract", "facility-employee-day-jobtype"),
    "WORK_HRS_NUM": ("Hours worked", "hours, two decimals", "facility-employee-day-jobtype"),
    "WORK_HRS_FN": ("Hours footnote", "file-specific flag/footnote field in current CMS API header", "facility-employee-day-jobtype"),
}


def build_time_period_policy_variable_map() -> list[dict[str, str]]:
    evidence = source_ev("CMS QSO-22-08-NH", "CMS July 27 2022 Care Compare update", "PBJ PUF April 2022 technical specifications", "Five-Star Technical Users' Guides Jan/Jul/Oct 2022 and Jan 2023")
    return [
        {
            "period_id": "2017Q1_to_2021Q4",
            "period_dates": "2017-01-01 to 2021-12-31",
            "policy_status": "PBJ daily public-use staffing raw data period before 2022 transparency shock; 2021Q4 has a documented incomplete-data exception for affected providers.",
            "public_reporting_status": "Standard staffing measures were public; weekend staffing and turnover were not yet newly posted as Care Compare measures.",
            "five_star_rating_status": "Staffing rating used pre-July method based on adjusted total nurse and adjusted RN staffing, with RN staffing rating reported separately.",
            "employee_level_data_status": "Employee-level detail not publicly available before the January 2022 release, even though underlying PBJ submissions existed.",
            "provider_catalog_field_status": "ProviderInfo contains adjusted total nurse HPRD, adjusted RN HPRD, staffing stars, and RN staffing rating; not the six-measure July 2022 point score.",
            "interpretation_for_pre_period": "This is the cleanest true pre-period for daily PBJ staffing behavior, excluding documented anomalies such as 2021Q4 ransomware truncation where needed.",
            "interpretation_for_research_design": "Use for baseline trends and pre-period outcomes; do not call January 2022 the first existence of PBJ daily staffing data.",
            "source_evidence": evidence,
        },
        {
            "period_id": "2022_01_07_to_2022_01_25",
            "period_dates": "2022-01-07 to 2022-01-25",
            "policy_status": "QSO announcement period; providers are informed that weekend staffing, turnover, and employee-level data will be public and later rating-relevant.",
            "public_reporting_status": "Announcement and anticipation; weekend/turnover public posting imminent.",
            "five_star_rating_status": "Pre-July rating method still in force.",
            "employee_level_data_status": "Employee-level detail announced but not yet posted.",
            "provider_catalog_field_status": "January Provider Data Catalog refresh begins to expose weekend and turnover facility-level measures.",
            "interpretation_for_pre_period": "Not a clean untreated pre-period because official announcement creates anticipation.",
            "interpretation_for_research_design": "Treat as announcement/anticipation window, not as untreated baseline.",
            "source_evidence": evidence,
        },
        {
            "period_id": "2022_01_26_to_2022_07_26",
            "period_dates": "2022-01-26 to 2022-07-26",
            "policy_status": "Transparency and employee-level data-release transition period before formal July staffing rating algorithm change.",
            "public_reporting_status": "Weekend staffing and turnover are public on Care Compare; employee-level PBJ detail is public.",
            "five_star_rating_status": "New six-measure staffing score not yet formally applied to the staffing-domain star rating.",
            "employee_level_data_status": "Public employee-level detail available; historical employee-detail files begin with 2020Q2 content.",
            "provider_catalog_field_status": "ProviderInfo contains reported weekend staffing and turnover fields, but not adjusted weekend total nurse HPRD or official 0-380 staffing score.",
            "interpretation_for_pre_period": "January-July 2022 is a transition/public-reporting/anticipation period, not a clean untreated pre-period.",
            "interpretation_for_research_design": "Useful for anticipation and transparency analyses; fragile as a pre-period for July RD-DID unless explicitly modeled.",
            "source_evidence": evidence,
        },
        {
            "period_id": "2022_07_27_to_2022_09_30",
            "period_dates": "2022-07-27 to 2022-09-30",
            "policy_status": "Formal staffing-domain rating algorithm event; six-measure point score determines staffing stars.",
            "public_reporting_status": "Weekend and turnover measures remain public; July release changes rating consequences.",
            "five_star_rating_status": "Six measures enter staffing rating; total score is mapped to 1-5 stars using 155/205/255/320 cutoffs.",
            "employee_level_data_status": "Employee-level detail remains public and is required conceptually for turnover measures.",
            "provider_catalog_field_status": "July ProviderInfo exposes reported weekend total nurse HPRD plus total adjustment components, turnover, and stars, but not the official 0-380 score or explicit adjusted weekend total nurse HPRD.",
            "interpretation_for_pre_period": "Not pre-period; this is the event window.",
            "interpretation_for_research_design": "Potential RD/RD-DID threshold event, but running variable is reconstructed and identification remains diagnostic-sensitive.",
            "source_evidence": evidence,
        },
        {
            "period_id": "2022Q4_onward",
            "period_dates": "2022-10-01 onward",
            "policy_status": "Post-algorithm period after transition quarter.",
            "public_reporting_status": "Weekend and turnover public reporting continues.",
            "five_star_rating_status": "Six-measure staffing rating method continues in technical guides and Provider Data Catalog.",
            "employee_level_data_status": "Employee-level detail continues to be public quarterly.",
            "provider_catalog_field_status": "October 2022 and January 2023 ProviderInfo expose explicit adjusted weekend total nurse HPRD in addition to other components.",
            "interpretation_for_pre_period": "Post-period only.",
            "interpretation_for_research_design": "Best post period for formula/label and metric-salience analyses; compare against true pre-period and isolate transition.",
            "source_evidence": evidence,
        },
        {
            "period_id": "2022_10_27_snapshot",
            "period_dates": "2022-10-27 Provider Data Catalog snapshot",
            "policy_status": "Post-July Provider Data Catalog snapshot.",
            "public_reporting_status": "Public ProviderInfo file includes the new facility-level measures.",
            "five_star_rating_status": "Six-measure method in force.",
            "employee_level_data_status": "Employee-level detail public.",
            "provider_catalog_field_status": "Explicit `Adjusted Weekend Total Nurse Staffing Hours per Resident per Day` field is present; official 0-380 staffing score is still not found.",
            "interpretation_for_pre_period": "Not pre-period.",
            "interpretation_for_research_design": "Validation snapshot for the adjusted-weekend reconstruction identity and star emulator.",
            "source_evidence": source_ev(evidence, "Provider Data Catalog archive nursing_homes_2022-10-27.zip"),
        },
        {
            "period_id": "2023_01_02_snapshot",
            "period_dates": "2023-01-02 Provider Data Catalog snapshot",
            "policy_status": "Post-July Provider Data Catalog snapshot.",
            "public_reporting_status": "Public ProviderInfo file includes the new facility-level measures.",
            "five_star_rating_status": "Six-measure method in force.",
            "employee_level_data_status": "Employee-level detail public.",
            "provider_catalog_field_status": "Explicit adjusted weekend total nurse HPRD remains present; official 0-380 staffing score is still not found.",
            "interpretation_for_pre_period": "Not pre-period.",
            "interpretation_for_research_design": "Second validation snapshot for reconstruction and star emulator.",
            "source_evidence": source_ev(evidence, "Provider Data Catalog archive nursing_homes_2023-01-02.zip"),
        },
    ]


def build_raw_pbj_daily_codebook() -> list[dict[str, str]]:
    data_file_name = "PBJ daily PUF official fields"
    try:
        sample = fetch_text_json(PBJ_DAILY_DATA, SNAP / "pbj_daily_sample_record.json")
        headers = list(sample["value"][0].keys())
        data_file_name = str(sample.get("value", [{}])[0].get("CY_Qtr", "current CMS API sample"))
    except Exception:
        headers = list(PBJ_DAILY_LABELS.keys())
    rows: list[dict[str, str]] = []
    evidence = source_ev("CMS PBJ Daily data-viewer headers", "PBJ PUF April 2022 technical specifications Table 1", "CMS PBJ staffing data submission page")
    for var in headers:
        desc, job, unit, freq = PBJ_DAILY_LABELS.get(var, (var, "", "", "facility-day"))
        rows.append(
            {
                "period_id": "2017Q1_to_current",
                "source_file_or_doc": "CMS PBJ Daily Nurse Staffing API data-viewer and PBJ PUF April 2022 technical specifications",
                "exact_raw_variable_name": var,
                "exact_label_or_description": desc,
                "job_code_if_applicable": job,
                "staff_type_if_applicable": staff_type_for_daily(var),
                "unit": unit,
                "frequency_or_unit_of_observation": freq,
                "first_available_period_verified": "2017Q1",
                "last_available_period_verified": "current CMS API/sample or PBJ PUF documentation: " + data_file_name,
                "appears_before_jan2022": "yes",
                "appears_after_jan2022": "yes",
                "notes": "Raw official PBJ daily nursing public-use field; not a project-constructed HPRD or outcome.",
                "source_evidence": evidence,
            }
        )
    rows.append(
        {
            "period_id": "2021Q4_exception",
            "source_file_or_doc": "PBJ PUF April 2022 technical specifications",
            "exact_raw_variable_name": "Incomplete",
            "exact_label_or_description": "Provider has incomplete staffing data for the 2021Q4 ransomware-affected subset",
            "job_code_if_applicable": "",
            "staff_type_if_applicable": "",
            "unit": "0/1 flag",
            "frequency_or_unit_of_observation": "facility-day in 2021Q4 PUF exception",
            "first_available_period_verified": "2021Q4",
            "last_available_period_verified": "2021Q4",
            "appears_before_jan2022": "yes",
            "appears_after_jan2022": "no",
            "notes": "Special exception field documented for 2021Q4 only; keep as data-quality flag, not a staffing outcome.",
            "source_evidence": source_ev("PBJ PUF April 2022 technical specifications lines on 2021Q4 incomplete field"),
        }
    )
    return rows


def staff_type_for_daily(var: str) -> str:
    if var in {"Hrs_RNDON", "Hrs_RNDON_emp", "Hrs_RNDON_ctr"}:
        return "RN director of nursing"
    if var in {"Hrs_RNadmin", "Hrs_RNadmin_emp", "Hrs_RNadmin_ctr"}:
        return "RN administrative"
    if var in {"Hrs_RN", "Hrs_RN_emp", "Hrs_RN_ctr"}:
        return "registered nurse"
    if var in {"Hrs_LPNadmin", "Hrs_LPNadmin_emp", "Hrs_LPNadmin_ctr"}:
        return "LPN administrative"
    if var in {"Hrs_LPN", "Hrs_LPN_emp", "Hrs_LPN_ctr"}:
        return "licensed practical/vocational nurse"
    if var in {"Hrs_CNA", "Hrs_CNA_emp", "Hrs_CNA_ctr"}:
        return "certified nurse aide"
    if var in {"Hrs_NAtrn", "Hrs_NAtrn_emp", "Hrs_NAtrn_ctr"}:
        return "nurse aide trainee"
    if var in {"Hrs_MedAide", "Hrs_MedAide_emp", "Hrs_MedAide_ctr"}:
        return "medication aide/technician"
    return ""


def build_pbj_employee_detail_codebook() -> list[dict[str, str]]:
    try:
        sample = fetch_text_json(PBJ_EMP_DATA, SNAP / "pbj_employee_detail_sample_record.json")
        headers = list(sample["value"][0].keys())
    except Exception:
        headers = list(EMPLOYEE_DETAIL_LABELS.keys())
    rows: list[dict[str, str]] = []
    evidence = source_ev("QSO-22-08-NH employee-level release on January 26, 2022", "PBJ Employee Detail Public Use File Technical Specifications January 2022", "CMS employee-detail data-viewer headers")
    for var in headers:
        desc, unit, freq = EMPLOYEE_DETAIL_LABELS.get(var, (var, "", "facility-employee-day-jobtype"))
        needed = "yes" if var in {"PROVNUM", "WorkDate", "SYS_EMPLEE_ID", "EMPLEE_JOB_CD_ID", "EMP_CTR", "WORK_HRS_NUM"} else "supporting"
        rows.append(
            {
                "period_id": "2022_01_26_release_and_after",
                "source_file_or_doc": "CMS PBJ Employee Detail API data-viewer and January 2022 Employee Detail PUF technical specifications",
                "exact_raw_variable_name": var,
                "exact_label_or_description": desc,
                "unit": unit,
                "frequency_or_unit_of_observation": freq,
                "first_public_availability": "January 26, 2022 public release; quarterly files include data beginning with 2020Q2 content",
                "needed_for_turnover": needed,
                "appears_before_jan2022_publicly": "no",
                "appears_after_jan2022_publicly": "yes",
                "notes": "Employee-level detail is needed to identify starts/stops/turnover; the public release changed availability, not the existence of underlying PBJ submissions.",
                "source_evidence": evidence,
            }
        )
    return rows


def build_provider_catalog_codebook(records: list[dict[str, Any]]) -> tuple[list[dict[str, str]], list[tuple[str, str]]]:
    latest = latest_theme_snapshot(records)
    specs = HISTORICAL_SNAPSHOT_SPECS + [(f"current_latest_{latest['date']}", latest["date"])]
    archive_paths: list[tuple[str, str]] = []
    target_fields = [
        ("facility identifier", "Federal Provider Number"),
        ("provider characteristics", "Provider Name"),
        ("provider characteristics", "Provider State"),
        ("provider characteristics", "Ownership Type"),
        ("provider characteristics", "Number of Certified Beds"),
        ("provider characteristics", "Average Number of Residents per Day"),
        ("reported total nurse HPRD", "Reported Total Nurse Staffing Hours per Resident per Day"),
        ("adjusted total nurse HPRD", "Adjusted Total Nurse Staffing Hours per Resident per Day"),
        ("reported RN HPRD", "Reported RN Staffing Hours per Resident per Day"),
        ("adjusted RN HPRD", "Adjusted RN Staffing Hours per Resident per Day"),
        ("reported weekend total nurse HPRD", "Total number of nurse staff hours per resident per day on the weekend"),
        ("reported weekend RN HPRD", "Registered Nurse hours per resident per day on the weekend"),
        ("adjusted weekend total nurse HPRD", "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day"),
        ("total nurse turnover", "Total nursing staff turnover"),
        ("RN turnover", "Registered Nurse turnover"),
        ("administrator turnover/departures", "Number of administrators who have left the nursing home"),
        ("staffing star rating", "Staffing Rating"),
        ("RN staffing rating", "RN Staffing Rating"),
        ("overall star rating", "Overall Rating"),
        ("health inspection rating", "Health Inspection Rating"),
        ("quality measure rating", "QM Rating"),
        ("unrelated score fields", "Rating Cycle 1 Total Health Score"),
        ("unrelated score fields", "Total Weighted Health Survey Score"),
    ]
    rows: list[dict[str, str]] = []
    for label, snap_date in specs:
        zp, source = provider_zip(snap_date, records)
        archive_paths.append((snap_date, source))
        member, header = read_provider_header(zp)
        dictionaries = read_provider_dictionary(zp)
        provider_desc = dictionaries.get("ProviderInfo", {})
        candidate_usecols = [f for _, f in target_fields if f in header]
        provider_df = read_provider_info(zp, candidate_usecols)
        for role, field in target_fields:
            present = field in header
            rows.append(
                {
                    "snapshot_date": snap_date,
                    "archive_zip_or_source": source,
                    "file_name": member,
                    "exact_raw_variable_name": field,
                    "exact_label_or_description": provider_desc.get(field, ""),
                    "normalized_project_name_if_any": norm_project_name(field),
                    "role": role,
                    "missingness": missingness(provider_df, field) if present else "field_absent",
                    "direct_official_field": "yes" if present else "no",
                    "constructed_in_project": "no",
                    "notes": provider_field_note(field, snap_date, present),
                    "source_evidence": source_ev(f"Provider Data Catalog archive {snap_date}", "NH_ProviderInfo CSV header", "NH_Primary_Data_Dictionary.xlsx ProviderInfo sheet"),
                }
            )
        for score in scan_score_fields(zp)[:40]:
            rows.append(
                {
                    "snapshot_date": snap_date,
                    "archive_zip_or_source": source,
                    "file_name": score["file_name"],
                    "exact_raw_variable_name": score["field"],
                    "exact_label_or_description": "Score field in a Provider Data Catalog table; audited as not the official facility-level staffing 0-380 score unless explicitly identified as staffing-domain score.",
                    "normalized_project_name_if_any": norm_project_name(score["field"]),
                    "role": score["role"],
                    "missingness": "not_computed_for_non_providerinfo_file",
                    "direct_official_field": "yes",
                    "constructed_in_project": "no",
                    "notes": "Included to prevent mistaking health/QM score fields for the missing official staffing point score.",
                    "source_evidence": source_ev(f"Provider Data Catalog archive {snap_date}", "CSV header score-field scan"),
                }
            )
    return rows, archive_paths


def provider_field_note(field: str, snapshot: str, present: bool) -> str:
    if field == "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day" and snapshot == "2022-07-27" and not present:
        return "Critical absence: July 2022 has reported weekend total nurse HPRD and all-day adjustment components, but not this explicit adjusted weekend field."
    if field == "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day" and present:
        return "Direct official field present; used to validate the July reconstruction identity."
    if field == "RN Staffing Rating" and not present and snapshot >= "2022-07-27":
        return "Separate RN staffing star field disappears after the July six-measure staffing-score reform in the audited snapshots."
    if present:
        return "Direct official Provider Data Catalog field in this snapshot."
    return "Expected/role-relevant field absent from this snapshot header."


def field_presence(provider_rows: list[dict[str, str]], snapshot: str, field: str) -> str:
    for row in provider_rows:
        if row["snapshot_date"] == snapshot and row["exact_raw_variable_name"] == field and row["file_name"].endswith(".csv"):
            return field if row["direct_official_field"] == "yes" else "absent"
    return "absent"


def build_six_measure_crosswalk(provider_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    snaps = {
        "jan": "2022-01-27",
        "apr": "2022-04-27",
        "jul": "2022-07-27",
        "oct": "2022-10-27",
        "jan2023": "2023-01-02",
    }
    measure_specs = [
        {
            "num": "1",
            "name": "Case-mix adjusted total nurse HPRD",
            "desc": "Case-mix adjusted total nursing hours per resident day for RN, LPN/LVN, and aide hours averaged across all days.",
            "role": "100-point staffing-level component; higher values receive more points.",
            "max": "100",
            "rule": "National deciles; lowest decile 10 points, highest decile 100 points.",
            "status": "case-mix adjusted official published ProviderInfo measure",
            "field": "Adjusted Total Nurse Staffing Hours per Resident per Day",
            "project": "adjusted_total_nurse_hprd",
            "direct": "direct official measure",
            "formula": "",
            "evidence_strength": "high_direct",
            "notes": "Directly safe as a component, subject to CMS exclusions/missingness.",
        },
        {
            "num": "2",
            "name": "Case-mix adjusted RN HPRD",
            "desc": "Case-mix adjusted RN hours per resident day averaged across all days.",
            "role": "100-point staffing-level component; higher values receive more points.",
            "max": "100",
            "rule": "National deciles; lowest decile 10 points, highest decile 100 points.",
            "status": "case-mix adjusted official published ProviderInfo measure",
            "field": "Adjusted RN Staffing Hours per Resident per Day",
            "project": "adjusted_rn_hprd",
            "direct": "direct official measure",
            "formula": "",
            "evidence_strength": "high_direct",
            "notes": "Directly safe as a component, subject to CMS exclusions/missingness.",
        },
        {
            "num": "3",
            "name": "Case-mix adjusted weekend total nurse HPRD",
            "desc": "Case-mix adjusted total nursing hours per resident day for RN, LPN/LVN, and aide hours averaged across weekend days.",
            "role": "50-point weekend staffing component; higher values receive more points.",
            "max": "50",
            "rule": "National deciles; lowest decile 5 points, highest decile 50 points.",
            "status": "adjusted official measure, but explicit July 2022 ProviderInfo field is absent",
            "field": "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day",
            "project": "adjusted_weekend_total_nurse_hprd_reconstructed",
            "direct": "reconstructed for July 2022; direct official field in October 2022 and January 2023",
            "formula": "reported_weekend_total_nurse_hprd * adjusted_total_nurse_hprd / reported_total_nurse_hprd",
            "evidence_strength": "strong_reconstruction_validated",
            "notes": "Safe enough to rerun the RD running variable as reconstructed official-like evidence, not a directly observed July score.",
        },
        {
            "num": "4",
            "name": "Total nurse turnover",
            "desc": "Percentage of nursing staff that left the nursing home over a twelve-month period.",
            "role": "50-point turnover component; lower turnover receives more points.",
            "max": "50",
            "rule": "National deciles; highest turnover decile 5 points, lowest turnover decile 50 points.",
            "status": "directly published official turnover measure derived from employee-level PBJ",
            "field": "Total nursing staff turnover",
            "project": "total_nursing_staff_turnover",
            "direct": "direct official measure",
            "formula": "CMS annual turnover method using six consecutive quarters of PBJ employee-level data",
            "evidence_strength": "high_direct",
            "notes": "Requires employee-level identifiers to calculate from raw data; direct facility-level measure is in ProviderInfo after January 2022.",
        },
        {
            "num": "5",
            "name": "RN turnover",
            "desc": "Percentage of RN staff that left the nursing home over a twelve-month period.",
            "role": "50-point turnover component; lower turnover receives more points.",
            "max": "50",
            "rule": "National deciles; highest turnover decile 5 points, lowest turnover decile 50 points.",
            "status": "directly published official turnover measure derived from employee-level PBJ",
            "field": "Registered Nurse turnover",
            "project": "rn_turnover",
            "direct": "direct official measure",
            "formula": "CMS annual turnover method using six consecutive quarters of PBJ employee-level data",
            "evidence_strength": "high_direct",
            "notes": "Requires employee-level identifiers to calculate from raw data; direct facility-level measure is in ProviderInfo after January 2022.",
        },
        {
            "num": "6",
            "name": "Administrator turnover / administrator departures",
            "desc": "Number of administrators who left the nursing home over a twelve-month period.",
            "role": "30-point administrator turnover component; fewer departures receive more points.",
            "max": "30",
            "rule": "0 departures = 30 points; 1 departure = 25 points; 2 or more departures = 10 points.",
            "status": "directly published official turnover/departure measure derived from employee-level PBJ",
            "field": "Number of administrators who have left the nursing home",
            "project": "administrator_departures",
            "direct": "direct official measure",
            "formula": "CMS administrator turnover/departure method using employee-level PBJ job code 1",
            "evidence_strength": "high_direct",
            "notes": "Direct facility-level measure is in ProviderInfo after January 2022; raw reconstruction requires employee-level job code and employee identifier.",
        },
    ]
    rows: list[dict[str, str]] = []
    for spec in measure_specs:
        field = spec["field"]
        rows.append(
            {
                "official_six_measure_number": spec["num"],
                "official_measure_name_from_technical_guide": spec["name"],
                "official_description": spec["desc"],
                "official_rating_role": spec["role"],
                "max_points": spec["max"],
                "point_rule": spec["rule"],
                "official_raw_or_adjusted_status": spec["status"],
                "exact_available_field_jan2022": available_field_text(provider_rows, snaps["jan"], field),
                "exact_available_field_apr2022": available_field_text(provider_rows, snaps["apr"], field),
                "exact_available_field_jul2022": available_field_text(provider_rows, snaps["jul"], field),
                "exact_available_field_oct2022": available_field_text(provider_rows, snaps["oct"], field),
                "exact_available_field_jan2023": available_field_text(provider_rows, snaps["jan2023"], field),
                "project_variable_name": spec["project"],
                "direct_or_reconstructed": spec["direct"],
                "reconstruction_formula_if_any": spec["formula"],
                "evidence_strength": spec["evidence_strength"],
                "source_evidence": source_ev("Five-Star Technical Users' Guide July 2022 staffing-domain section and Table 3/Table A2", "Provider Data Catalog ProviderInfo field audit", "V3 emulator validation table"),
                "notes": spec["notes"],
            }
        )
    return rows


def available_field_text(provider_rows: list[dict[str, str]], snapshot: str, field: str) -> str:
    present = field_presence(provider_rows, snapshot, field)
    if present != "absent":
        return field
    if field == "Adjusted Weekend Total Nurse Staffing Hours per Resident per Day" and snapshot in {"2022-01-27", "2022-04-27", "2022-07-27"}:
        return "absent; reported field available: Total number of nurse staff hours per resident per day on the weekend"
    return "absent"


def build_constructed_variables_codebook() -> list[dict[str, str]]:
    ev_pbj = source_ev("PBJ Daily raw headers", "PBJ PUF April 2022 technical specifications", "project V2/V3 scripts")
    ev_guide = source_ev("Five-Star Technical Users' Guide July 2022", "Provider Data Catalog field audit", "V3 score rescue scripts and emulator validation")
    rows = [
        cv("rn_hprd", "(Hrs_RNDON + Hrs_RNadmin + Hrs_RN) / MDScensus", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "hours per resident day", "constructed from raw PBJ daily fields", "2017Q1", "yes", "yes", "yes", "yes_as_component_or_outcome", "no", "Daily/quarterly RN staffing intensity; not resident clinical quality.", ev_pbj),
        cv("total_nurse_hprd", "(Hrs_RNDON + Hrs_RNadmin + Hrs_RN + Hrs_LPNadmin + Hrs_LPN + Hrs_CNA + Hrs_NAtrn + Hrs_MedAide) / MDScensus", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; Hrs_LPNadmin; Hrs_LPN; Hrs_CNA; Hrs_NAtrn; Hrs_MedAide; MDScensus", "hours per resident day", "constructed from raw PBJ daily fields", "2017Q1", "yes", "yes", "yes", "yes_as_component_or_outcome", "no", "Research-created daily/period HPRD unless using CMS published adjusted ProviderInfo measure.", ev_pbj),
        cv("weekend_total_nurse_hprd", "sum(total nurse hours on Saturdays/Sundays) / sum(MDScensus on Saturdays/Sundays)", "WorkDate; nurse-hour fields; MDScensus", "hours per resident day", "constructed from raw PBJ daily fields; also has reported official ProviderInfo measure after Jan 2022", "2017Q1", "yes", "yes", "yes", "yes_component_input", "no", "Reported weekend total is directly public from Jan 2022, but raw construction is available earlier.", ev_pbj),
        cv("weekend_rn_hprd", "sum(RN hours on Saturdays/Sundays) / sum(MDScensus on Saturdays/Sundays)", "WorkDate; Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "hours per resident day", "constructed from raw PBJ daily fields; also has reported official ProviderInfo measure after Jan 2022", "2017Q1", "yes", "yes", "yes", "no_not_in_six_component_score", "no", "Public transparency measure, but not one of the July six score components.", ev_pbj),
        cv("rn_lt8h_day_indicator_share", "mean(1[RN hours < 8 and MDScensus > 0])", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "share of days", "purely research-created outcome", "2017Q1", "yes", "yes", "yes", "no", "no", "Primary staffing reliability outcome; not a clinical quality outcome.", ev_pbj),
        cv("zero_rn_day_indicator_share", "mean(1[RN hours == 0 and MDScensus > 0])", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "share of days", "purely research-created outcome", "2017Q1", "yes", "yes", "yes", "no", "no", "Primary reliability/safety staffing measure; treat 2021Q4 carefully.", ev_pbj),
        cv("weekend_p10_hprd", "10th percentile of weekend total nurse HPRD within facility-period", "WorkDate; nurse-hour fields; MDScensus", "hours per resident day", "purely research-created outcome", "2017Q1", "yes", "yes", "yes", "no", "no", "Lower-tail staffing reliability outcome.", ev_pbj),
        cv("weekend_p25_hprd", "25th percentile of weekend total nurse HPRD within facility-period", "WorkDate; nurse-hour fields; MDScensus", "hours per resident day", "purely research-created outcome", "2017Q1", "yes", "yes", "yes", "no", "no", "Lower-tail staffing reliability outcome.", ev_pbj),
        cv("worst_weekend_hprd", "minimum weekend total nurse HPRD within facility-period", "WorkDate; nurse-hour fields; MDScensus", "hours per resident day", "purely research-created outcome", "2017Q1", "yes", "yes", "yes", "no", "no", "Lower-tail staffing reliability outcome; sensitive to extreme reporting days.", ev_pbj),
        cv("contract_share", "contract nurse hours / total nurse hours", "all *_ctr nursing fields; all total nursing hour fields", "share", "purely research-created staffing mix variable", "2017Q1", "yes", "yes", "yes", "no", "possible_exposure_or_outcome", "Staffing mix/reallocation measure, not a CMS rating component.", ev_pbj),
        cv("weekend_share_of_total_hours", "weekend nurse hours / all nurse hours in period", "WorkDate; nurse-hour fields", "share", "purely research-created staffing allocation variable", "2017Q1", "yes", "yes", "yes", "no", "possible_exposure_or_outcome", "Metric-salience/reallocation outcome.", ev_pbj),
        cv("reconstructed_adjusted_weekend_total_nurse_hprd", "reported_weekend_total_nurse_hprd * adjusted_total_nurse_hprd / reported_total_nurse_hprd", "Total number of nurse staff hours per resident per day on the weekend; Adjusted Total Nurse Staffing Hours per Resident per Day; Reported Total Nurse Staffing Hours per Resident per Day", "hours per resident day", "reconstructed official-like measure", "2022-07-27", "no_for_true_pre_jan_public_fields; yes_from_raw PBJ if independently constructed", "yes", "no", "yes_component", "no", "Used because July adjusted weekend official field is absent; validated against Oct 2022 and Jan 2023 explicit fields.", ev_guide),
        cv("reconstructed_staffing_score_0_380", "sum six component points, rescaled to 380 if CMS component missingness rules require", "six official staffing components", "0-380 points", "reconstructed official-like score", "2022-07-27", "no_old_method_pre_july", "yes", "no", "yes_running_variable", "no", "Official score not found directly; reconstruction validated to July staffing stars at 0.963.", ev_guide),
        cv("staffing_star_threshold_indicators", "1[reconstructed_staffing_score >= 155/205/255/320]", "reconstructed_staffing_score_0_380", "binary indicators", "project-created threshold exposure from reconstructed official-like score", "2022-07-27", "no", "yes", "no", "yes_running_variable", "yes", "Used for RD/RD-DID threshold designs; inherits running-variable limitations.", ev_guide),
        cv("formula_induced_overall_star_loss", "indicator for losing old four-star staffing bonus under July overall-star rule while holding observed domain ratings fixed", "overall_rating; health_inspection_rating; staffing_rating; qm_rating; July and Jan 2022 formula rules", "binary indicator", "project-created formula-shock exposure", "2022-07-27", "no", "yes", "no", "no", "yes", "Mechanism/exposure variable; exclusion requires diagnostics and should not be read as pure randomized treatment.", ev_guide),
        cv("high_shadow_price_rating_incentive", "indicator for facilities near a staffing threshold or with large marginal rating/overall-rating payoff", "reconstructed_staffing_score_0_380; staffing star thresholds; formula-induced overall-star loss", "binary/score", "project-created incentive variable", "2022-07-27", "no", "yes", "no", "yes_context", "yes", "Mechanism/heterogeneity variable for label incentive, not a raw official field.", ev_guide),
    ]
    return rows


def cv(name: str, formula: str, inputs: str, unit: str, source_type: str, first: str, pre: str, post: str, outcome: str, rv: str, exposure: str, notes: str, ev: str) -> dict[str, str]:
    return {
        "constructed_variable_name": name,
        "formula": formula,
        "inputs_exact_raw_field_names": inputs,
        "unit": unit,
        "constructed_from_raw_or_official_component": source_type,
        "first_period_constructible": first,
        "valid_pre_period": pre,
        "valid_post_period": post,
        "used_as_outcome": outcome,
        "used_as_running_variable_component": rv,
        "used_as_treatment_or_exposure": exposure,
        "notes": notes,
        "source_evidence": ev,
    }


def build_running_variable_audit() -> list[dict[str, str]]:
    validation_path = V3_ROOT / "result" / "tables" / "emulator_validation_v3.csv"
    if validation_path.exists():
        val = pd.read_csv(validation_path)
        july = float(val.loc[val["snapshot_date"].eq("2022-07-27"), "star_match_rate"].iloc[0])
        octv = float(val.loc[val["snapshot_date"].eq("2022-10-27"), "star_match_rate"].iloc[0])
        janv = float(val.loc[val["snapshot_date"].eq("2023-01-02"), "star_match_rate"].iloc[0])
        match = f"July 2022={july:.3f}; Oct 2022={octv:.3f}; Jan 2023={janv:.3f}"
    else:
        match = "not available"
    ev = source_ev("Provider Data Catalog field hunt", "Five-Star Technical Users' Guide July 2022", "V3 emulator_validation_v3.csv")
    return [
        {
            "running_variable_name": "official_facility_level_staffing_score_0_380",
            "official_or_reconstructed": "official field searched, not found",
            "source_components": "No direct field found in ProviderInfo headers, data dictionaries, or score-field scan.",
            "formula": "",
            "official_cutoffs": "155, 205, 255, 320",
            "cutoffs_source": "Five-Star Technical Users' Guide July 2022 Table 3",
            "validation_target": "Official staffing star rating",
            "validation_match_rate": "not applicable because direct official score field is absent",
            "validation_period": "January 2022, April 2022, July 2022, October 2022, January 2023 Provider Data Catalog snapshots",
            "validity_status": "not_directly_observed",
            "limitations": "The official 0-380 score cannot be treated as directly observed in public ProviderInfo files.",
            "design_implication": "RD can only use a reconstructed official-like score, not a direct official running variable.",
            "source_evidence": ev,
        },
        {
            "running_variable_name": "reconstructed_staffing_score_0_380",
            "official_or_reconstructed": "reconstructed official-like score",
            "source_components": "Adjusted total nurse HPRD; adjusted RN HPRD; reconstructed adjusted weekend total nurse HPRD; total nurse turnover; RN turnover; administrator departures.",
            "formula": "Assign official Table A2 points for six measures, sum to 380 maximum, and apply CMS rescaling/rounding when component missingness requires.",
            "official_cutoffs": "155, 205, 255, 320",
            "cutoffs_source": "Five-Star Technical Users' Guide July 2022 Table 3",
            "validation_target": "Official staffing star rating",
            "validation_match_rate": match,
            "validation_period": "Primary July 2022 validation; October 2022 and January 2023 used as explicit adjusted-weekend checks.",
            "validity_status": "candidate_validated_reconstruction",
            "limitations": "High star-match supports rerunning RD/RD-DID but does not prove local random assignment, no sorting, covariate balance, or strong causal identification.",
            "design_implication": "Use as a transparent reconstructed running variable; report density, balance, pre-outcome, and placebo diagnostics before causal claims.",
            "source_evidence": ev,
        },
    ]


def build_outcome_family_codebook() -> list[dict[str, str]]:
    ev = source_ev("PBJ Daily raw fields", "Provider Data Catalog ProviderInfo fields", "project V2/V3 outcome scripts")
    return [
        outcome("Staffing behavior / staffing reliability", "zero_rn_day_share", "constructed", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "2017Q1 onward", "facility-day aggregated to facility-period", "Share of resident days/weekend days with no RN hours.", "close staffing reliability", "yes", "Sensitive to reporting artifacts and 2021Q4 truncation; not resident clinical quality.", ev),
        outcome("Staffing behavior / staffing reliability", "rn_lt8h_day_share", "constructed", "Hrs_RNDON; Hrs_RNadmin; Hrs_RN; MDScensus", "2017Q1 onward", "facility-day aggregated to facility-period", "Share of days with RN hours below eight.", "close staffing reliability", "yes", "Threshold is research-defined and should be justified.", ev),
        outcome("Staffing behavior / staffing reliability", "weekend_p10_total_hprd", "constructed", "weekend total nurse hours; MDScensus", "2017Q1 onward", "facility-weekend/day aggregated", "Lower-tail weekend staffing intensity.", "directly targeted by July weekend measure", "yes", "Extreme values need cleaning rules.", ev),
        outcome("Staffing behavior / staffing reliability", "worst_weekend_total_hprd", "constructed", "weekend total nurse hours; MDScensus", "2017Q1 onward", "facility-period", "Minimum weekend staffing intensity.", "directly targeted by July weekend measure", "yes", "Can be noisy; pair with percentile outcomes.", ev),
        outcome("Staffing mix / contract labor / reallocation", "contract_share", "constructed", "nursing *_ctr fields; total nursing hours", "2017Q1 onward", "facility-day or facility-period", "Reliance on contract labor among nursing hours.", "staffing mix not direct rating score", "no", "Useful mechanism, not primary quality outcome.", ev),
        outcome("Staffing mix / contract labor / reallocation", "weekend_share_of_total_hours", "constructed", "WorkDate; total nurse hours", "2017Q1 onward", "facility-period", "Allocation of nurse hours toward weekends.", "metric-salience/reallocation", "secondary", "Interpret with total hours to distinguish reallocation from expansion.", ev),
        outcome("Demand / operation", "avg_daily_census", "raw/constructed from MDScensus", "MDScensus", "2017Q1 onward", "facility-day aggregated", "Resident census / demand scale.", "not direct staffing policy outcome", "no", "Can be affected by occupancy/demand shocks.", ev),
        outcome("Ratings", "staffing_rating", "direct official ProviderInfo field", "Staffing Rating", "Provider snapshots", "facility-snapshot", "Public staffing star label.", "direct policy label", "secondary", "Rating outcome is partly mechanical under the July formula.", ev),
        outcome("Ratings", "overall_rating", "direct official ProviderInfo field", "Overall Rating", "Provider snapshots", "facility-snapshot", "Overall public star label.", "formula-mediated public label", "secondary", "Formula-induced changes are labels, not clinical quality.", ev),
        outcome("Ratings", "rn_staffing_rating", "direct official ProviderInfo field where present", "RN Staffing Rating", "pre-July Provider snapshots; absent in audited July onward snapshots", "facility-snapshot", "Separate RN staffing star under old method.", "old-method rating context", "secondary", "Not the July six-measure score.", ev),
        outcome("Deficiencies / downstream quality", "health_deficiency_score", "direct official ProviderInfo field", "Rating Cycle 1/2/3 Total Health Score; Total Weighted Health Survey Score", "Provider snapshots", "facility-snapshot", "Inspection/survey deficiency score.", "downstream/secondary", "no_primary", "Not immediate staffing behavior; timing and survey cycles complicate causal attribution.", ev),
        outcome("Deficiencies / downstream quality", "quality_measure_rating", "direct official ProviderInfo field", "QM Rating", "Provider snapshots", "facility-snapshot", "Quality measure star domain.", "secondary public quality domain", "no_primary", "Do not describe as resident clinical quality improvement without separate clinical outcome evidence.", ev),
    ]


def outcome(family: str, name: str, raw: str, inputs: str, period: str, level: str, interp: str, proximity: str, primary: str, limits: str, ev: str) -> dict[str, str]:
    return {
        "outcome_family": family,
        "outcome_variable_name": name,
        "raw_or_constructed": raw,
        "exact_inputs": inputs,
        "period_available": period,
        "level_of_observation": level,
        "interpretation": interp,
        "policy_proximity": proximity,
        "should_be_primary": primary,
        "limitations": limits,
        "source_evidence": ev,
    }


def md_table(rows: list[dict[str, str]], columns: list[str], max_rows: int | None = None) -> str:
    show = rows if max_rows is None else rows[:max_rows]
    lines = ["| " + " | ".join(columns) + " |", "| " + " | ".join(["---"] * len(columns)) + " |"]
    for row in show:
        vals = []
        for c in columns:
            val = str(row.get(c, "")).replace("\n", " ").replace("|", "/")
            vals.append(val[:260])
        lines.append("| " + " | ".join(vals) + " |")
    if max_rows is not None and len(rows) > max_rows:
        lines.append("")
        lines.append(f"Showing {max_rows} of {len(rows)} rows.")
    return "\n".join(lines)


def build_report(
    time_rows: list[dict[str, str]],
    daily_rows: list[dict[str, str]],
    employee_rows: list[dict[str, str]],
    provider_rows: list[dict[str, str]],
    six_rows: list[dict[str, str]],
    constructed_rows: list[dict[str, str]],
    running_rows: list[dict[str, str]],
    outcome_rows: list[dict[str, str]],
    latest_snapshot: str,
) -> str:
    clean_table_cols = ["period_id", "period_dates", "public_reporting_status", "five_star_rating_status", "provider_catalog_field_status", "interpretation_for_research_design"]
    provider_summary = [
        r
        for r in provider_rows
        if r["role"]
        in {
            "reported weekend total nurse HPRD",
            "adjusted weekend total nurse HPRD",
            "total nurse turnover",
            "staffing star rating",
            "RN staffing rating",
        }
        and r["file_name"].endswith(".csv")
    ]
    lines = [
        "# Final Time-Versioned Codebook",
        "",
        "## Executive Summary",
        "",
        "This metadata audit rebuilds the nursing-home staffing project around time-versioned variables before any new causal modeling. The corrected reading is straightforward: PBJ daily nurse staffing raw data existed before 2022, with quarterly public-use files beginning in 2017Q1. January 2022 is not the first existence of PBJ daily staffing data; it is the transparency and employee-level release moment. July 27, 2022 is the formal staffing-domain rating algorithm event, when the six-measure point score entered Five-Star staffing ratings.",
        "",
        "The public Provider Data Catalog does not expose a direct facility-level official 0-380 staffing score in the audited snapshots. July 2022 also does not expose an explicit `Adjusted Weekend Total Nurse Staffing Hours per Resident per Day` field. The project can reconstruct that adjusted weekend component from official July fields using the all-day case-mix adjustment ratio and can reconstruct a 0-380 official-like staffing score; V3 validation matched July 2022 official staffing stars at 0.963. That supports rerunning threshold designs, but it does not by itself establish strong causal identification.",
        "",
        "## Corrected Conceptual Understanding",
        "",
        "This is not a simple DiD plus RD project. The credible design language is RD-DID / threshold design around July 2022 staffing-score cutoffs, formula-induced label shock from the overall-star rule change, and metric-salience DDD comparing targeted weekend staffing metrics with less-targeted staffing dimensions.",
        "",
        "January-July 2022 is a transition/public-reporting/anticipation period, not a clean untreated pre-period. The cleanest pre-period for staffing behavior is 2017Q1-2021Q4, with special handling for documented 2021Q4 PBJ incompleteness. July 2022 is the formal algorithmic rating event. Ratings and deficiencies should stay secondary; staffing reliability outcomes are the primary behavior outcomes. Do not call staffing/rating changes resident clinical quality improvement.",
        "",
        "## Timeline",
        "",
        md_table(time_rows, clean_table_cols),
        "",
        "## Raw PBJ Daily Variables By Period",
        "",
        "Raw PBJ daily fields are facility-day official fields. They are separate from HPRD, weekend shares, low-tail reliability measures, and reconstructed score components.",
        "",
        md_table(daily_rows, ["exact_raw_variable_name", "exact_label_or_description", "job_code_if_applicable", "staff_type_if_applicable", "unit", "first_available_period_verified", "appears_before_jan2022", "appears_after_jan2022"], max_rows=40),
        "",
        "## Employee-Level PBJ Variables And Why January 2022 Matters",
        "",
        "Employee-level PBJ data were not publicly available before January 26, 2022. The public employee-detail file contains facility, state, quarter, work date, system employee identifier, job code, employee/contract type, and hours. Turnover measures require employee-level identifiers, job codes, dates, and hours because starts/stops/departures cannot be recovered from facility-day aggregate PBJ daily fields alone.",
        "",
        md_table(employee_rows, ["exact_raw_variable_name", "exact_label_or_description", "unit", "first_public_availability", "needed_for_turnover", "appears_before_jan2022_publicly", "appears_after_jan2022_publicly"]),
        "",
        "## Provider Data Catalog Variables By Snapshot",
        "",
        f"The latest official nursing-homes archive snapshot observed from the CMS archive API during this audit is {latest_snapshot}. ProviderInfo field availability was checked variable-by-snapshot for January 2022, April 2022, July 2022, October 2022, January 2023, and that current/latest snapshot.",
        "",
        md_table(provider_summary, ["snapshot_date", "exact_raw_variable_name", "role", "missingness", "direct_official_field", "notes"], max_rows=80),
        "",
        "## The Six Official Staffing-Rating Measures",
        "",
        md_table(six_rows, ["official_six_measure_number", "official_measure_name_from_technical_guide", "max_points", "point_rule", "exact_available_field_jul2022", "direct_or_reconstructed", "project_variable_name", "evidence_strength", "notes"]),
        "",
        "## What Exists Directly, What Must Be Reconstructed, And What Is Unavailable",
        "",
        "Raw variables: PBJ daily fields such as `Hrs_RN`, `Hrs_LPN`, `Hrs_CNA`, `MDScensus`, `WorkDate`, and employee/contract splits are official raw fields. Official published measures: ProviderInfo fields such as adjusted total nurse HPRD, adjusted RN HPRD, reported weekend total nurse HPRD, turnover measures, and star ratings are direct official measures when present. Reconstructed official-like variables: July adjusted weekend total nurse HPRD and the 0-380 staffing score are reconstructed from official components. Research-created outcomes: RN<8h days, zero-RN days, lower-tail weekend HPRD, contract share, and reallocation metrics are project-created outcomes.",
        "",
        md_table(constructed_rows, ["constructed_variable_name", "constructed_from_raw_or_official_component", "first_period_constructible", "valid_pre_period", "valid_post_period", "used_as_outcome", "used_as_running_variable_component"], max_rows=30),
        "",
        "## Running Variable Explanation",
        "",
        md_table(running_rows, ["running_variable_name", "official_or_reconstructed", "source_components", "official_cutoffs", "validation_match_rate", "validity_status", "design_implication"]),
        "",
        "The official facility-level 0-380 staffing score is not directly found. The RD running variable is a reconstructed official-like 0-380 score based on the six official components and Table A2 point rules. The July adjusted weekend component is reconstructed as `reported_weekend_total_nurse_hprd * adjusted_total_nurse_hprd / reported_total_nurse_hprd`, supported by the technical-guide case-mix adjustment logic and checked against later snapshots where the adjusted weekend field is explicit.",
        "",
        "## Outcome Explanation",
        "",
        md_table(outcome_rows, ["outcome_family", "outcome_variable_name", "raw_or_constructed", "period_available", "interpretation", "policy_proximity", "should_be_primary", "limitations"], max_rows=40),
        "",
        "## Implications For Causal Design",
        "",
        "- True pre-period: use 2017Q1-2021Q4 PBJ daily staffing behavior, with 2021Q4 data-quality sensitivity checks.",
        "- Transition period: treat January 7/26 through July 26, 2022 as public-reporting, employee-level release, and anticipation rather than clean untreated baseline.",
        "- July 2022 RD/RD-DID: possible with the reconstructed running variable, but fragile. It still needs density, covariate balance, pre-outcome, bandwidth, and placebo-cutoff diagnostics.",
        "- Old high/low exposure DiD: remains weak because broad exposure definitions do not solve pretrend and timing concerns.",
        "- Formula-induced label shock: useful as conditional mechanism evidence because it isolates rating-label consequences of the July overall-star rule, but exclusion remains fragile.",
        "- Metric-salience DDD: useful for asking whether facilities changed targeted weekend staffing metrics more than less-targeted staffing dimensions.",
        "",
        "## Final Clean Table For Human Readers",
        "",
        md_table(time_rows, clean_table_cols),
        "",
        "## Source-Grounded Audit Files",
        "",
        "Machine-readable outputs are in `result/tables/`. The build script is `script/build_metadata_audit.py`; the formal self-check is `script/self_check_metadata_audit.py`.",
    ]
    return "\n".join(lines)


def main() -> None:
    ensure_dirs()
    records = archive_api_records()
    latest = latest_theme_snapshot(records)
    latest_snapshot = str(latest["date"])

    time_rows = build_time_period_policy_variable_map()
    daily_rows = build_raw_pbj_daily_codebook()
    employee_rows = build_pbj_employee_detail_codebook()
    provider_rows, archive_paths = build_provider_catalog_codebook(records)
    six_rows = build_six_measure_crosswalk(provider_rows)
    constructed_rows = build_constructed_variables_codebook()
    running_rows = build_running_variable_audit()
    outcome_rows = build_outcome_family_codebook()

    write_csv(TABLES / "time_period_policy_variable_map.csv", time_rows)
    write_csv(TABLES / "raw_pbj_daily_codebook_by_period.csv", daily_rows)
    write_csv(TABLES / "pbj_employee_detail_codebook_by_period.csv", employee_rows)
    write_csv(TABLES / "provider_catalog_codebook_by_snapshot.csv", provider_rows)
    write_csv(TABLES / "six_measure_rating_component_crosswalk.csv", six_rows)
    write_csv(TABLES / "constructed_variables_codebook.csv", constructed_rows)
    write_csv(TABLES / "running_variable_audit.csv", running_rows)
    write_csv(TABLES / "outcome_family_codebook.csv", outcome_rows)

    write_csv(
        AUDIT_LOGS / "provider_archive_sources_used.csv",
        [{"snapshot_date": d, "archive_zip_or_source": s} for d, s in archive_paths],
    )
    report = build_report(time_rows, daily_rows, employee_rows, provider_rows, six_rows, constructed_rows, running_rows, outcome_rows, latest_snapshot)
    write_text(REPORT / "final_time_versioned_codebook.md", report)


if __name__ == "__main__":
    main()
